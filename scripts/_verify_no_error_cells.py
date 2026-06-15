"""Prove the demo workbook is clean.

openpyxl stores formulas as strings and never evaluates them, so a *freshly written*
file can't surface a live ``#NAME?`` the way Excel would. The #NAME? defect we just
fixed was prose accidentally written into a cell starting with "=" — Excel parses the
sentence as a formula and yields #NAME?. So this verifier runs two scans:

  (1) the literal scan the task asked for — every cell value starting with "#"
      (#NAME?/#REF!/#VALUE!/#DIV0!…); must be zero.
  (2) the substantive scan that catches the actual defect class — every *formula* cell
      (data_type 'f') must parse as a plausible Excel formula, never an English
      sentence. We strip quoted string-literals, then flag any formula whose remaining
      body contains a space (real formulas here have spaces only inside quoted text) or
      a leading implicit-intersection "@". Either signals prose-as-formula.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ERROR_PREFIXES = ("#NAME?", "#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!", "#")
_QUOTED = re.compile(r'"[^"]*"')


def scan(path: Path) -> int:
    wb = load_workbook(path)
    hash_hits: list[str] = []
    prose_formula_hits: list[str] = []

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                ref = f"{ws.title}!{cell.coordinate}"
                if v.startswith("#"):
                    hash_hits.append(f"{ref}: {v!r}")
                if cell.data_type == "f" or v.startswith("="):
                    body = _QUOTED.sub("", v[1:] if v.startswith("=") else v)
                    if v.startswith("=@") or " " in body:
                        prose_formula_hits.append(f"{ref}: {v!r}")

    n_cells = sum(ws.max_row * ws.max_column for ws in wb.worksheets)
    print(f"Workbook: {path}")
    print(f"Tabs ({len(wb.worksheets)}): {', '.join(ws.title for ws in wb.worksheets)}")
    print(f"Cells scanned (bounding box): ~{n_cells}")
    print(f"(1) cells with value starting '#' : {len(hash_hits)}")
    for h in hash_hits:
        print(f"      {h}")
    print(f"(2) prose-as-formula cells        : {len(prose_formula_hits)}")
    for h in prose_formula_hits:
        print(f"      {h}")

    bad = len(hash_hits) + len(prose_formula_hits)
    print("RESULT:", "CLEAN" if bad == 0 else f"DIRTY ({bad} problem cells)")
    return bad


if __name__ == "__main__":
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("deal_model_demo.xlsx")
    sys.exit(1 if scan(target.resolve()) else 0)
