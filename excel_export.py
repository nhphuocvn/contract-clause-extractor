"""Export contract extraction JSON outputs to a multi-sheet Excel workbook.

Usage:
    python excel_export.py sample_contract.json sample_contract_2.json -o Portfolio.xlsx

The workbook has four sheets: README, Portfolio, AI Risk Findings, Source Quotes.
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Style palette
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill("solid", fgColor="F2F2F2")
BODY_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=18, color="1F3864")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="1F3864")

SEVERITY_FILLS = {
    "high":               PatternFill("solid", fgColor="F8696B"),
    "medium":             PatternFill("solid", fgColor="FFEB84"),
    "low":                PatternFill("solid", fgColor="C6EFCE"),
    "low_protection":     PatternFill("solid", fgColor="FFEB84"),
    "review_recommended": PatternFill("solid", fgColor="FFEB84"),
    "informational":      PatternFill("solid", fgColor="F2F2F2"),
}

_thin = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(vertical="center", horizontal="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Defensive JSON accessors
# ---------------------------------------------------------------------------


def _safe_dict(obj) -> dict:
    return obj if isinstance(obj, dict) else {}


def field_value(extraction: dict, key: str) -> str:
    f = _safe_dict(extraction.get(key))
    v = f.get("value", "")
    return "" if v is None else str(v)


def field_quote(extraction: dict, key: str) -> str:
    f = _safe_dict(extraction.get(key))
    q = f.get("source_quote", "")
    return "" if q is None else str(q)


def sub_value(extraction: dict, parent: str, child: str) -> str:
    p = _safe_dict(extraction.get(parent))
    c = _safe_dict(p.get(child))
    v = c.get("value", "")
    return "" if v is None else str(v)


def sub_quote(extraction: dict, parent: str, child: str) -> str:
    p = _safe_dict(extraction.get(parent))
    c = _safe_dict(p.get(child))
    q = c.get("source_quote", "")
    return "" if q is None else str(q)


def parties_string(extraction: dict) -> str:
    parties = _safe_dict(extraction.get("parties"))
    v = parties.get("value", []) or []
    if not isinstance(v, list):
        return ""
    return "; ".join(
        f"{p.get('name', '') or ''} ({p.get('role', '') or ''})".strip()
        for p in v if isinstance(p, dict)
    )


def parties_quote(extraction: dict) -> str:
    return field_quote(extraction, "parties")


def contract_type_primary(extraction: dict) -> str:
    ct = _safe_dict(extraction.get("contract_type"))
    return str(ct.get("primary", "") or "")


def ai_overall_risk(extraction: dict) -> str:
    ai = _safe_dict(extraction.get("ai_risk_assessment"))
    return str(ai.get("overall_risk_level", "") or "")


def deterministic_overall(extraction: dict) -> str:
    rs = _safe_dict(extraction.get("risk_summary"))
    return str(rs.get("overall_risk", "") or "")


def ai_findings(extraction: dict) -> list[dict]:
    ai = _safe_dict(extraction.get("ai_risk_assessment"))
    fs = ai.get("findings", []) or []
    return [f for f in fs if isinstance(f, dict)]


def risk_fill_for(level: str):
    return SEVERITY_FILLS.get((level or "").lower())


# ---------------------------------------------------------------------------
# Sheet helpers
# ---------------------------------------------------------------------------


def write_header(ws, headers, row=1):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 24
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate


def style_data_row(ws, row_idx, alt=False):
    for cell in ws[row_idx]:
        cell.font = BODY_FONT
        cell.alignment = WRAP_TOP
        cell.border = THIN_BORDER
        # Don't overwrite a severity-colored fill that was set earlier.
        if alt and cell.fill.fgColor.value in (None, "00000000", "FFFFFFFF"):
            cell.fill = ALT_ROW_FILL


def set_column_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def setup_print(ws, landscape=True):
    ws.page_setup.orientation = (
        ws.ORIENTATION_LANDSCAPE if landscape else ws.ORIENTATION_PORTRAIT
    )
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.gridLines = False


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def build_readme_sheet(ws, contracts_meta: list[tuple[str, str]]):
    """contracts_meta: list of (source_filename, contract_type_primary)."""
    setup_print(ws, landscape=False)
    set_column_widths(ws, [30, 88])

    ws["A1"] = "Contract Portfolio Export"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = BODY_FONT
    ws.merge_cells("A2:B2")

    row = 4
    ws.cell(row=row, column=1, value="Purpose").font = SUBTITLE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    purpose_cell = ws.cell(
        row=row, column=1,
        value=(
            "This workbook summarizes outputs of the Contract Clause Extractor pipeline. "
            "Each contract is parsed into 12 structured fields with verbatim source quotes, "
            "classified by type, and assessed for risk against company standard terms. The "
            "audit-trail sheet pairs every extracted value with the contract excerpt it was "
            "taken from."
        ),
    )
    purpose_cell.alignment = WRAP_TOP
    purpose_cell.font = BODY_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.row_dimensions[row].height = 64

    row += 2
    ws.cell(row=row, column=1, value="Sheets").font = SUBTITLE_FONT
    row += 1
    for name, desc in [
        ("README", "This sheet — purpose, pipeline notes, contracts in this run."),
        ("Portfolio", "One row per contract with key terms and overall risk level."),
        ("AI Risk Findings", "Flattened paralegal findings across all contracts."),
        ("Source Quotes", "Audit trail pairing each extracted value with its verbatim contract excerpt."),
    ]:
        c1 = ws.cell(row=row, column=1, value=name)
        c1.font = BOLD_FONT
        c1.alignment = WRAP_TOP
        c2 = ws.cell(row=row, column=2, value=desc)
        c2.alignment = WRAP_TOP
        c2.font = BODY_FONT
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Source pipeline").font = SUBTITLE_FONT
    row += 1
    for line in [
        "1. extract.py reads each contract (.txt or .pdf) and calls OpenAI gpt-4o-mini with structured output.",
        "2. The model returns 12 fields with verbatim source quotes, a contract-type classification, and an AI risk assessment that compares against company standard terms.",
        "3. A deterministic rule check runs alongside as a cross-check and fallback if the AI assessment fails.",
        "4. excel_export.py compiles the JSON outputs from step 1-3 into this workbook.",
    ]:
        c = ws.cell(row=row, column=1, value=line)
        c.alignment = WRAP_TOP
        c.font = BODY_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1
    title_cell = ws.cell(row=row, column=1, value=f"Contracts in this run ({len(contracts_meta)})")
    title_cell.font = SUBTITLE_FONT
    row += 1
    for src, ctype in contracts_meta:
        c1 = ws.cell(row=row, column=1, value=src)
        c1.font = BOLD_FONT
        c2 = ws.cell(row=row, column=2, value=ctype or "—")
        c2.font = BODY_FONT
        c2.alignment = WRAP_TOP
        row += 1


PORTFOLIO_HEADERS = [
    "Source file", "Contract type", "Parties", "Effective date", "Term length",
    "Auto-renewal", "Payment terms", "SLA", "Liability cap", "Indemnity cap",
    "Governing law", "Confidentiality", "AI overall risk", "Deterministic risk",
]

PORTFOLIO_WIDTHS = [28, 26, 36, 16, 18, 32, 32, 22, 26, 26, 22, 18, 16, 18]


def build_portfolio_sheet(ws, contracts: list[tuple[str, dict]]):
    setup_print(ws, landscape=True)
    write_header(ws, PORTFOLIO_HEADERS)
    set_column_widths(ws, PORTFOLIO_WIDTHS)

    for i, (src, data) in enumerate(contracts):
        row_idx = i + 2
        values = [
            src,
            contract_type_primary(data),
            parties_string(data),
            field_value(data, "effective_date"),
            field_value(data, "term_length"),
            field_value(data, "auto_renewal"),
            field_value(data, "payment_terms"),
            field_value(data, "sla_commitments"),
            field_value(data, "limitation_of_liability"),
            field_value(data, "indemnity_cap"),
            field_value(data, "governing_law"),
            field_value(data, "confidentiality_period"),
            ai_overall_risk(data) or "—",
            deterministic_overall(data) or "—",
        ]
        for col, v in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col, value=v)

        # Risk-column coloring (cols 13 and 14)
        ai_fill = risk_fill_for(ai_overall_risk(data))
        if ai_fill:
            ws.cell(row=row_idx, column=13).fill = ai_fill
        det_fill = risk_fill_for(deterministic_overall(data))
        if det_fill:
            ws.cell(row=row_idx, column=14).fill = det_fill

        style_data_row(ws, row_idx, alt=(i % 2 == 1))
        ws.row_dimensions[row_idx].height = 60


FINDINGS_HEADERS = [
    "Source file", "Severity", "Category", "Title",
    "Finding", "Standard deviation", "Counter-position",
]

FINDINGS_WIDTHS = [28, 14, 22, 36, 60, 50, 50]


def build_findings_sheet(ws, contracts: list[tuple[str, dict]]):
    setup_print(ws, landscape=True)
    write_header(ws, FINDINGS_HEADERS)
    set_column_widths(ws, FINDINGS_WIDTHS)

    row_idx = 2
    band_toggle = 0
    for src, data in contracts:
        findings = ai_findings(data)
        if not findings:
            continue
        for f in findings:
            severity = str(f.get("severity", "") or "")
            ws.cell(row=row_idx, column=1, value=src)
            sev_cell = ws.cell(row=row_idx, column=2, value=severity)
            ws.cell(row=row_idx, column=3, value=str(f.get("category", "") or ""))
            ws.cell(row=row_idx, column=4, value=str(f.get("title", "") or ""))
            ws.cell(row=row_idx, column=5, value=str(f.get("finding", "") or ""))
            ws.cell(row=row_idx, column=6, value=str(f.get("standard_deviation", "") or ""))
            ws.cell(row=row_idx, column=7, value=str(f.get("counter_position", "") or ""))

            sev_fill = risk_fill_for(severity)
            if sev_fill:
                sev_cell.fill = sev_fill

            style_data_row(ws, row_idx, alt=(band_toggle % 2 == 1))
            ws.row_dimensions[row_idx].height = 70
            row_idx += 1
            band_toggle += 1

    if row_idx == 2:
        # No findings at all — leave a friendly note.
        ws.cell(row=2, column=1, value="No AI findings to report across the contracts in this run.")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(FINDINGS_HEADERS))
        ws.cell(row=2, column=1).font = BODY_FONT
        ws.cell(row=2, column=1).alignment = WRAP_TOP


QUOTES_HEADERS = ["Source file", "Field", "Extracted value", "Verbatim source quote"]
QUOTES_WIDTHS = [28, 36, 50, 80]


FIELD_ORDER = [
    ("parties", "parties"),
    ("effective_date", "effective_date"),
    ("term_length", "term_length"),
    ("auto_renewal", "auto_renewal"),
    ("payment_terms", "payment_terms"),
    ("termination_clauses.for_cause", ("termination_clauses", "for_cause")),
    ("termination_clauses.for_convenience", ("termination_clauses", "for_convenience")),
    ("termination_clauses.for_non_payment", ("termination_clauses", "for_non_payment")),
    ("sla_commitments", "sla_commitments"),
    ("indemnity_cap", "indemnity_cap"),
    ("limitation_of_liability", "limitation_of_liability"),
    ("governing_law", "governing_law"),
    ("confidentiality_period", "confidentiality_period"),
    ("data_protection.encryption", ("data_protection", "encryption")),
    ("data_protection.data_residency", ("data_protection", "data_residency")),
    ("data_protection.certifications", ("data_protection", "certifications")),
    ("data_protection.compliance_frameworks", ("data_protection", "compliance_frameworks")),
]


def _resolve_field(extraction: dict, accessor):
    """Return (value_str, quote_str) for a top-level or nested field accessor."""
    if accessor == "parties":
        return parties_string(extraction), parties_quote(extraction)
    if isinstance(accessor, tuple):
        parent, child = accessor
        return sub_value(extraction, parent, child), sub_quote(extraction, parent, child)
    return field_value(extraction, accessor), field_quote(extraction, accessor)


def build_quotes_sheet(ws, contracts: list[tuple[str, dict]]):
    setup_print(ws, landscape=True)
    write_header(ws, QUOTES_HEADERS)
    set_column_widths(ws, QUOTES_WIDTHS)

    row_idx = 2
    band_toggle = 0
    for src, data in contracts:
        for label, accessor in FIELD_ORDER:
            value, quote = _resolve_field(data, accessor)
            if not value and not quote:
                continue
            ws.cell(row=row_idx, column=1, value=src)
            ws.cell(row=row_idx, column=2, value=label)
            ws.cell(row=row_idx, column=3, value=value)
            ws.cell(row=row_idx, column=4, value=quote)
            style_data_row(ws, row_idx, alt=(band_toggle % 2 == 1))
            ws.row_dimensions[row_idx].height = 56
            row_idx += 1
            band_toggle += 1


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def load_contracts(paths: list[Path]) -> list[tuple[str, dict]]:
    contracts = []
    for p in paths:
        if not p.exists():
            print(f"[skip] {p}: file not found", file=sys.stderr)
            continue
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except Exception as exc:
            print(f"[skip] {p}: invalid JSON ({exc})", file=sys.stderr)
            continue
        if not isinstance(data, dict):
            print(f"[skip] {p}: top-level JSON is not an object", file=sys.stderr)
            continue
        contracts.append((p.name, data))
    return contracts


def build_workbook(contracts: list[tuple[str, dict]], output_path: Path) -> None:
    wb = Workbook()
    # Default sheet → README
    ws_readme = wb.active
    ws_readme.title = "README"

    contracts_meta = [(src, contract_type_primary(data)) for src, data in contracts]

    build_readme_sheet(ws_readme, contracts_meta)
    build_portfolio_sheet(wb.create_sheet("Portfolio"), contracts)
    build_findings_sheet(wb.create_sheet("AI Risk Findings"), contracts)
    build_quotes_sheet(wb.create_sheet("Source Quotes"), contracts)

    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Export contract extraction JSONs to a multi-sheet Excel workbook."
    )
    parser.add_argument("files", nargs="+", help="One or more *.json outputs from extract.py")
    parser.add_argument(
        "-o", "--output",
        default="Contract_Portfolio.xlsx",
        help="Output .xlsx filename (default: Contract_Portfolio.xlsx)",
    )
    args = parser.parse_args()

    paths = [Path(f) for f in args.files]
    contracts = load_contracts(paths)
    if not contracts:
        print("No valid contract JSONs to export.", file=sys.stderr)
        return 1

    output_path = Path(args.output)
    build_workbook(contracts, output_path)

    print(f"Wrote {output_path} with {len(contracts)} contract(s):")
    for src, data in contracts:
        ctype = contract_type_primary(data) or "(type unknown)"
        ai = ai_overall_risk(data) or "—"
        det = deterministic_overall(data) or "—"
        n_findings = len(ai_findings(data))
        print(f"  - {src}: {ctype} | AI risk: {ai} | det risk: {det} | findings: {n_findings}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
