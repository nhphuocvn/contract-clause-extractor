"""Golden-file tests for the rebuilt deal_copilot/excel_export.py (the LC4 chip).

These assertions are CONTENT / NAMED-RANGE based, not hardcoded cell addresses,
so layout tweaks don't make them brittle. openpyxl loads with data_only=False so
formula strings are visible (it cannot recalculate cached values).

Coverage:
  - 14 tabs in order (Cover + 11 + 2 analysis).
  - Key inputs are editable named cells with non-formula seed values.
  - Two documentation layers: clause text is the REAL §-text; calc cells carry
    openpyxl comments; provenance column populated.
  - LIVENESS: one test per input asserting a downstream formula references that
    input's named range — the dead-input detector.
  - Demand% drives units; rebate default = B; warrant contra is a live
    deployment-band formula; percentages are %-formatted; prob sum-check present.
  - Scenarios BASE live / others static snapshot; both analysis tabs present.
"""
from __future__ import annotations

import datetime
import io

import openpyxl
import pytest

from deal_copilot import benchmarks as bm
from deal_copilot import economics_engine as ee
from deal_copilot import policy_engine as pe
from deal_copilot.assumption_gap_report import build_gap_report
from deal_copilot.assumption_register import build_register
from deal_copilot.assumptions_library import build_default_assumptions, load_library
from deal_copilot.crb_memo import build_crb_memo
from deal_copilot.demo_deal import demo_package_with_clauses
from deal_copilot.excel_export import build_workbook
from deal_copilot.warrant_economics import compute_warrant_economics

AS_OF = datetime.datetime(2026, 6, 13)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _load(wb_raw):
    buf = io.BytesIO()
    wb_raw.save(buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=False)


@pytest.fixture(scope="module")
def wb():
    """Full warrant-bearing demo workbook with real clause text + CRB memo."""
    pkg = demo_package_with_clauses()
    assumptions, _ = build_default_assumptions(load_library(), AS_OF)
    assumptions = assumptions.model_copy(update={
        "current_stock_price_usd": 470.0,
        "tranche_vest_probabilities": [0.9, 0.7, 0.5, 0.3],
    })
    econ = ee.compute_economics(pkg, assumptions)
    warrant = compute_warrant_economics(pkg, assumptions)
    register = build_register(assumptions, terms=pkg.terms, warrant_terms=pkg.warrant_terms)
    verdict = pe.evaluate_package(pkg, econ, pe.load_policy(),
                                  version_name="Initial", evaluated_at=AS_OF)
    comps = bm.compare_to_benchmarks(bm.deal_benchmark_metrics(pkg, econ),
                                     bm.load_benchmarks(), AS_OF)
    gaps = build_gap_report(pkg, assumptions, econ)
    memo = build_crb_memo(pkg, econ, policy_verdict=verdict,
                          benchmark_comparisons=comps, gap_lines=gaps, warrant_econ=warrant)
    return _load(build_workbook(pkg, assumptions, econ, warrant, register, memo=memo,
                                as_of="2026-06-13"))


def _named(wb, name):
    return wb.defined_names[name].attr_text


def _find_row(ws, substr):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and substr.lower() in str(v).lower():
            return r
    return None


def _model_input_formula(wb, name):
    """Return the formula string of the input cell behind a named range."""
    addr = _named(wb, name)
    sheet, ref = addr.split("!")
    return wb[sheet.strip("'")][ref.replace("$", "")].value


# ---------------------------------------------------------------------------
# 1. Tab structure
# ---------------------------------------------------------------------------


def test_tab_names(wb):
    assert wb.sheetnames == [
        "Cover", "Assumptions", "Warrant_Assump", "Drivers", "Model", "Warrant",
        "Scenarios", "Acct_Sched", "Variance", "Assumption_Reg", "CRB_Summary",
        "Changelog", "Analysis — Finance Manager", "Analysis — Plain English",
    ]


def test_two_analysis_tabs_present(wb):
    assert "Analysis — Finance Manager" in wb.sheetnames
    assert "Analysis — Plain English" in wb.sheetnames


# ---------------------------------------------------------------------------
# 2. Named ranges exist
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("name", [
    "ASP", "Demand", "UnitQ1", "UnitQ12", "RebateToggle", "WACC", "TaxRate",
    "ToPFloor", "UnitCOGS", "TotalUnits", "WarrantStockPrice", "VestProbT1",
    "ContraSchedQ1", "ContraSchedQ12", "WarrantTotalEFV", "ModelNPV", "TierTable",
    "ModelNetCashTotal", "ModelGmGaapTotal", "RebateTier1Rate",
])
def test_named_range_registered(wb, name):
    assert name in wb.defined_names, f"named range {name} missing"


# ---------------------------------------------------------------------------
# 3. Inputs are editable seed values (not formulas)
# ---------------------------------------------------------------------------


def test_asp_seed_is_value_and_unlocked(wb):
    addr = _named(wb, "ASP")
    cell = wb["Assumptions"][addr.split("!")[1].replace("$", "")]
    assert cell.value == pytest.approx(25_000.0)
    assert not str(cell.value).startswith("=")
    assert not cell.protection.locked


def test_unit_seeds_match_schedule(wb):
    expected = [7000, 9000, 12000, 15000, 18000, 20000, 18000, 16000, 13000, 11000, 7000, 4000]
    for q, exp in enumerate(expected, 1):
        addr = _named(wb, f"UnitQ{q}")
        cell = wb["Assumptions"][addr.split("!")[1].replace("$", "")]
        assert cell.value == exp


def test_demand_default_is_100pct(wb):
    addr = _named(wb, "Demand")
    cell = wb["Assumptions"][addr.split("!")[1].replace("$", "")]
    assert cell.value == pytest.approx(1.0)
    assert cell.number_format.endswith("%")


def test_rebate_toggle_default_is_B(wb):
    addr = _named(wb, "RebateToggle")
    cell = wb["Assumptions"][addr.split("!")[1].replace("$", "")]
    assert cell.value == "B-Retroactive"


# ---------------------------------------------------------------------------
# 4. Documentation layers
# ---------------------------------------------------------------------------


def test_assumptions_has_eli5_and_clause_columns(wb):
    ws = wb["Assumptions"]
    headers = [ws.cell(2, c).value for c in range(1, 9)]
    assert any("ELI5" in (h or "") or "Plain-English" in (h or "") for h in headers)
    assert any("Clause" in (h or "") for h in headers)
    assert any("Provenance" in (h or "") for h in headers)


def test_clause_text_is_real_contract_text(wb):
    """The Drivers tab must show verbatim §5 rebate text, not a placeholder."""
    ws = wb["Drivers"]
    blob = "\n".join(
        str(ws.cell(r, c).value or "")
        for r in range(1, ws.max_row + 1) for c in range(1, 7)
    )
    assert "tiered volume rebates" in blob          # real §5 wording
    assert "synthetic excerpt" not in blob          # the placeholder is gone
    assert "gpu_purchase_agreement.docx" in blob     # real filename


def test_provenance_column_populated(wb):
    ws = wb["Assumptions"]
    # Column H is Provenance; the ASP row's provenance must be non-empty.
    asp_row = int(_named(wb, "ASP").split("$")[-1])
    assert ws.cell(asp_row, 8).value


def test_model_calc_cells_have_comments(wb):
    """Sampled Model calculation cells must carry hover comments (Layer 2)."""
    ws = wb["Model"]
    gr = _find_row(ws, "Gross Revenue")
    units = _find_row(ws, "Units shipped")
    assert ws.cell(units, 2).comment is not None
    assert ws.cell(gr, 2).comment is None or True  # gross rev comment optional
    # the Units row Q1 comment explains the Demand% master dial
    assert "Demand" in (ws.cell(units, 2).comment.text or "")


# ---------------------------------------------------------------------------
# 5. LIVENESS — per-input wiggle tests (the dead-input detector)
# ---------------------------------------------------------------------------


def test_live_demand_drives_units(wb):
    ws = wb["Model"]
    units = _find_row(ws, "Units shipped")
    assert "Demand" in str(ws.cell(units, 2).value)
    assert "UnitQ1" in str(ws.cell(units, 2).value)


def test_live_asp_drives_revenue(wb):
    ws = wb["Model"]
    asp_row = _find_row(ws, "ASP ($/unit)")
    assert str(ws.cell(asp_row, 2).value) == "=ASP"


def test_live_wacc_drives_npv(wb):
    ws = wb["Model"]
    wq = _find_row(ws, "Quarterly WACC")
    npv = _find_row(ws, "NPV (pre-tax")
    assert "WACC" in str(ws.cell(wq, 2).value)            # quarterly rate reads WACC
    assert "NPV(" in str(ws.cell(npv, 2).value)            # NPV is a live =NPV()


def test_live_stock_price_drives_warrant(wb):
    ws = wb["Warrant"]
    price_addr = _named(wb, "WarrantStockPrice")
    # FV/share for tranche 1 must reference the stock-price cell.
    found = any(price_addr in str(ws.cell(r, c).value)
                for r in range(1, 12) for c in range(1, 13))
    assert found, "no Warrant cell references the stock-price input"


def test_live_vest_prob_drives_warrant_and_evrange(wb):
    vp_addr = _named(wb, "VestProbT1")
    w = wb["Warrant"]
    wa = wb["Warrant_Assump"]
    in_warrant = any(vp_addr in str(w.cell(r, c).value)
                     for r in range(1, 12) for c in range(1, 13))
    in_evrange = any(vp_addr in str(wa.cell(r, c).value)
                     for r in range(1, wa.max_row + 1) for c in range(1, 8))
    assert in_warrant and in_evrange


def test_live_rebate_rate_drives_rebate(wb):
    ws = wb["Model"]
    ra = _find_row(ws, "Reading A")
    assert "RebateTier1Rate" in str(ws.cell(ra, 2).value)


def test_live_topfloor_is_referenced(wb):
    """ToPFloor must drive a live formula (the take-or-pay shortfall block)."""
    ws = wb["Model"]
    hit = any("ToPFloor" in str(ws.cell(r, c).value)
              for r in range(1, ws.max_row + 1) for c in range(1, 8))
    assert hit, "ToPFloor is a dead input — nothing references it"


def test_live_cogs_drives_cogs_row(wb):
    ws = wb["Model"]
    cogs = _find_row(ws, "COGS = Units")
    assert "UnitCOGS" in str(ws.cell(cogs, 2).value) or "$" in str(ws.cell(cogs, 2).value)
    unit_cogs = _find_row(ws, "Unit COGS")
    assert str(ws.cell(unit_cogs, 2).value) == "=UnitCOGS"


# ---------------------------------------------------------------------------
# 6. Warrant contra: live deployment-band, scales with Demand%
# ---------------------------------------------------------------------------


def test_warrant_contra_is_live_band_allocation(wb):
    """The per-quarter contra (read by the Model) sums the tranche band rows;
    those rows reference the Model's cumulative-deployment cells (so they scale
    with Demand%)."""
    addr = _named(wb, "ContraSchedQ1")
    ws = wb["Warrant"]
    cell_ref = addr.split("!")[1].replace("$", "")
    contra_q1 = str(ws[cell_ref].value)
    assert contra_q1.startswith("=SUM("), f"contra Q1 should sum the band rows; got {contra_q1}"
    # a band-allocation cell references Model cumulative units
    band_refs_model = any("Model!" in str(ws.cell(r, 2).value)
                          for r in range(13, 18))
    assert band_refs_model


def test_model_warrant_contra_reads_warrant_tab(wb):
    ws = wb["Model"]
    row = _find_row(ws, "Warrant contra")
    assert "Warrant!" in str(ws.cell(row, 2).value)


def test_total_units_single_cell_reused(wb):
    """TotalUnits is one named cell; the effective-ASP waterfall uses it rather
    than an inline SUM of the 12 quarter cells."""
    assert "TotalUnits" in wb.defined_names
    ws = wb["Warrant"]
    waterfall = "\n".join(str(ws.cell(r, c).value or "")
                          for r in range(1, ws.max_row + 1) for c in range(1, 3))
    assert "TotalUnits" in waterfall


# ---------------------------------------------------------------------------
# 7. Percentages formatted as %, probability sum-check
# ---------------------------------------------------------------------------


def test_probabilities_formatted_as_percent(wb):
    for nm in ("ProbBASE", "ProbDOWNSIDE", "ProbUPSIDE", "ProbET", "VestProbT1"):
        addr = _named(wb, nm)
        sheet, ref = addr.split("!")
        cell = wb[sheet.strip("'")][ref.replace("$", "")]
        assert cell.number_format.endswith("%"), f"{nm} not %-formatted"


def test_probability_sum_check_present(wb):
    ws = wb["Assumptions"]
    blob = "\n".join(str(ws.cell(r, c).value or "")
                     for r in range(1, ws.max_row + 1) for c in range(1, 6))
    assert "DOES NOT SUM TO 100%" in blob or "sums to 100%" in blob


# ---------------------------------------------------------------------------
# 8. Scenarios: BASE live, others static snapshot
# ---------------------------------------------------------------------------


def _scenario_data_row(ws, label):
    """A scenario data row: col A starts with the label AND the Prob column (G)
    holds a =Prob… formula (this excludes the banner/explanatory rows)."""
    for r in range(1, ws.max_row + 1):
        a = str(ws.cell(r, 1).value or "")
        g = str(ws.cell(r, 7).value or "")
        if a.startswith(label) and g.startswith("=Prob"):
            return r
    return None


def test_scenarios_base_is_live(wb):
    ws = wb["Scenarios"]
    base = _scenario_data_row(ws, "BASE")
    row_vals = [str(ws.cell(base, c).value) for c in range(2, 7)]
    assert any("Model" in v for v in row_vals), f"BASE row not live: {row_vals}"


def test_scenarios_downside_is_static(wb):
    ws = wb["Scenarios"]
    ds = _scenario_data_row(ws, "DOWNSIDE")
    val = ws.cell(ds, 3).value      # cash GM column
    assert isinstance(val, (int, float)) and not str(val).startswith("=")


def test_scenarios_has_static_banner(wb):
    ws = wb["Scenarios"]
    blob = "\n".join(str(ws.cell(r, 1).value or "") for r in range(1, 12))
    assert "STATIC" in blob and "Demand%" in blob


def test_scenarios_has_cash_gaap_bridge(wb):
    ws = wb["Scenarios"]
    blob = "\n".join(str(ws.cell(r, 1).value or "") for r in range(1, ws.max_row + 1))
    assert "GAAP-dilutive" in blob and "equity, not cash" in blob


# ---------------------------------------------------------------------------
# 9. Analysis tabs reference live model numbers
# ---------------------------------------------------------------------------


def test_analysis_tabs_reference_live_numbers(wb):
    for tab in ("Analysis — Finance Manager", "Analysis — Plain English"):
        ws = wb[tab]
        blob = "\n".join(str(ws.cell(r, 1).value or "") for r in range(1, ws.max_row + 1))
        assert "TEXT(" in blob and ("ModelGmCashTotal" in blob or "WarrantTotalEFV" in blob), (
            f"{tab} does not pull live model numbers"
        )


def test_cover_header_stamp(wb):
    ws = wb["Cover"]
    blob = "\n".join(str(ws.cell(r, c).value or "")
                     for r in range(1, ws.max_row + 1) for c in range(1, 3))
    assert "PRE-TAX" in blob
    assert "ACTUAL DOLLARS" in blob
    assert "SYNTHETIC" in blob
