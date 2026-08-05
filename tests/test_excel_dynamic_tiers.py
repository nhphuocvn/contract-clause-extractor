"""Stage 1 — the Model tab's rebate block must GENERATE from the tier count.

The engine (`driver_mapper._normalize_tiers`) and the Assumptions tab have always
handled N rebate tiers; only the Model tab was pinned at exactly 3. These tests
build the same deal at 5, 2 and 3 tiers and assert the headroom rows, zone rows,
named ranges and Reading-A formula all follow the actual tier count — and that
the rows below the rebate block shift instead of colliding.
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from deal_copilot import economics_engine as ee
from deal_copilot.assumption_register import build_register
from deal_copilot.excel_export import build_workbook
from deal_copilot.schemas import TermType
from deal_copilot.warrant_economics import compute_warrant_economics
from tests.fixtures import synthetic_package_with_warrant, warrant_assumptions

AS_OF = "2026-06-13"

TIERS_5 = [(20_000, 0.02), (40_000, 0.03), (70_000, 0.04), (100_000, 0.055), (130_000, 0.07)]
TIERS_3 = [(30_000, 0.03), (75_000, 0.05), (120_000, 0.07)]
TIERS_2 = [(50_000, 0.04), (110_000, 0.06)]


def _package_with_tiers(tiers):
    pkg = synthetic_package_with_warrant()
    for t in pkg.terms:
        if t.term_type == TermType.REBATE:
            t.parameters["tiers"] = [
                {"threshold_cumulative_units": thr, "pct_off_base_asp": rate}
                for thr, rate in tiers
            ]
    return pkg


def _build(tiers):
    pkg = _package_with_tiers(tiers)
    assumptions = warrant_assumptions()
    econ = ee.compute_economics(pkg, assumptions)
    warrant = compute_warrant_economics(pkg, assumptions)
    register = build_register(assumptions, terms=pkg.terms, warrant_terms=pkg.warrant_terms)
    wb = build_workbook(pkg, assumptions, econ, warrant, register, as_of=AS_OF)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=False)


@pytest.fixture(scope="module")
def wb5():
    return _build(TIERS_5)


@pytest.fixture(scope="module")
def wb3():
    return _build(TIERS_3)


@pytest.fixture(scope="module")
def wb2():
    return _build(TIERS_2)


def _labels(ws):
    return [str(ws.cell(r, 1).value or "") for r in range(1, ws.max_row + 1)]


def _rows_matching(ws, predicate):
    return [r for r in range(1, ws.max_row + 1) if predicate(str(ws.cell(r, 1).value or ""))]


def _find_row(ws, substr):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and substr.lower() in str(v).lower():
            return r
    return None


# ---------------------------------------------------------------------------
# 5 tiers
# ---------------------------------------------------------------------------


def test_five_tiers_render_five_headroom_rows(wb5):
    ws = wb5["Model"]
    heads = _rows_matching(ws, lambda s: s.startswith("Headroom below Tier"))
    assert len(heads) == 5
    assert heads == list(range(heads[0], heads[0] + 5)), "headroom rows must be contiguous"
    labels = [str(ws.cell(r, 1).value) for r in heads]
    assert labels == [f"Headroom below Tier {i}" for i in range(1, 6)]


def test_five_tiers_render_six_zone_rows(wb5):
    ws = wb5["Model"]
    zones = _rows_matching(ws, lambda s: s.startswith("Units in ") and s.endswith("zone"))
    assert len(zones) == 6, "N tiers must produce N+1 zones"
    labels = [str(ws.cell(r, 1).value) for r in zones]
    assert labels[0] == "Units in no-rebate zone"
    assert labels[1:] == [f"Units in Tier-{i} zone" for i in range(1, 6)]


def test_five_tiers_named_ranges_exist(wb5):
    for i in range(1, 6):
        assert f"RebateTier{i}Rate" in wb5.defined_names
        assert f"RebateTier{i}Threshold" in wb5.defined_names
    assert "RebateTier6Rate" not in wb5.defined_names


def test_five_tiers_reading_a_references_every_tier_rate(wb5):
    ws = wb5["Model"]
    ra = _find_row(ws, "Reading A")
    formula = str(ws.cell(ra, 2).value)
    for i in range(1, 6):
        assert f"RebateTier{i}Rate" in formula, f"Reading A ignores tier {i}: {formula}"
    assert "RebateTier6Rate" not in formula


def test_five_tiers_active_rebate_still_toggles_both_readings(wb5):
    ws = wb5["Model"]
    ra, rb = _find_row(ws, "Reading A"), _find_row(ws, "Reading B")
    act = _find_row(ws, "Active Rebate")
    formula = str(ws.cell(act, 2).value)
    assert "RebateToggle" in formula
    assert f"$B{rb}" in formula and f"$B{ra}" in formula


def test_five_tiers_reading_b_vlookup_path_intact(wb5):
    ws = wb5["Model"]
    rb = _find_row(ws, "Reading B")
    assert "VLOOKUP" in str(ws.cell(rb, 2).value)
    assert "TierTable" in str(ws.cell(rb, 2).value)
    # the tier table itself carries a zero row plus one row per tier
    start, end = wb5.defined_names["TierTable"].attr_text.split("!")[1].split(":")
    first = int(start.replace("$", "")[1:])
    last = int(end.replace("$", "")[1:])
    assert last - first + 1 == 6


def test_five_tiers_zone_rows_reference_matching_headroom_rows(wb5):
    """Zone i must be bounded by headroom i and headroom i-1 — no stale offsets."""
    ws = wb5["Model"]
    heads = _rows_matching(ws, lambda s: s.startswith("Headroom below Tier"))
    zones = _rows_matching(ws, lambda s: s.startswith("Units in ") and s.endswith("zone"))
    assert str(ws.cell(zones[0], 2).value) == f"=MIN($B4,$B{heads[0]})"
    for i in range(1, 5):
        assert str(ws.cell(zones[i], 2).value) == (
            f"=MIN($B4,$B{heads[i]})-MIN($B4,$B{heads[i - 1]})")
    assert str(ws.cell(zones[5], 2).value) == f"=$B4-MIN($B4,$B{heads[-1]})"


def test_five_tiers_rows_below_rebate_shift_without_collision(wb5, wb3):
    """Two extra tiers = four extra rows (2 headroom + 2 zones); everything below
    the rebate block must move down by exactly that much and stay unique."""
    m5, m3 = wb5["Model"], wb3["Model"]
    for label in ("Net Revenue — GAAP", "Gross Margin — cash", "Operating cash flow",
                  "NPV (pre-tax"):
        assert _find_row(m5, label) == _find_row(m3, label) + 4, f"{label} did not shift"
    # no two calculation rows landed on the same row
    labels = [s for s in _labels(m5) if s and not s.startswith("—")]
    assert len(labels) == len(set(labels)), "duplicate row labels imply a collision"


def test_five_tiers_downstream_tabs_follow_the_shift(wb5):
    """Acct_Sched reads the Model's active-rebate and prepayment rows by number."""
    m5, acct = wb5["Model"], wb5["Acct_Sched"]
    act = _find_row(m5, "Active Rebate")
    blob = "\n".join(str(acct.cell(r, c).value or "")
                     for r in range(1, acct.max_row + 1) for c in range(1, 9))
    assert f"Model!$B${act}" in blob, "rebate accrual points at the wrong Model row"
    for label in ("Prepayment available", "Prepayment drawdown", "Prepayment remaining"):
        assert f"Model!$B${_find_row(m5, label)}" in blob


def test_five_tiers_warrant_contra_still_reads_cumulative_deployment(wb5):
    m5, w5 = wb5["Model"], wb5["Warrant"]
    cum_start = _find_row(m5, "Cumulative units (start")
    blob = "\n".join(str(w5.cell(r, c).value or "")
                     for r in range(1, w5.max_row + 1) for c in range(1, 14))
    assert f"Model!$B${cum_start}" in blob


# ---------------------------------------------------------------------------
# 2 tiers
# ---------------------------------------------------------------------------


def test_two_tiers_render_two_headroom_and_three_zones(wb2):
    ws = wb2["Model"]
    heads = _rows_matching(ws, lambda s: s.startswith("Headroom below Tier"))
    zones = _rows_matching(ws, lambda s: s.startswith("Units in ") and s.endswith("zone"))
    assert len(heads) == 2
    assert len(zones) == 3
    assert str(ws.cell(zones[-1], 1).value) == "Units in Tier-2 zone"


def test_two_tiers_named_ranges_and_reading_a(wb2):
    for i in (1, 2):
        assert f"RebateTier{i}Rate" in wb2.defined_names
        assert f"RebateTier{i}Threshold" in wb2.defined_names
    assert "RebateTier3Rate" not in wb2.defined_names
    ws = wb2["Model"]
    formula = str(ws.cell(_find_row(ws, "Reading A"), 2).value)
    assert "RebateTier1Rate" in formula and "RebateTier2Rate" in formula
    assert "RebateTier3Rate" not in formula


def test_two_tiers_rows_shift_up_relative_to_three(wb2, wb3):
    m2, m3 = wb2["Model"], wb3["Model"]
    for label in ("Net Revenue — GAAP", "NPV (pre-tax"):
        assert _find_row(m2, label) == _find_row(m3, label) - 2


# ---------------------------------------------------------------------------
# 3 tiers — no regression against the shipped layout
# ---------------------------------------------------------------------------


def test_three_tiers_match_the_established_layout(wb3):
    """The standard deal must land on exactly the rows the shipped model used."""
    ws = wb3["Model"]
    expected = {
        "Units shipped": 4,
        "ASP ($/unit)": 5,
        "Gross Revenue": 6,
        "Cumulative units (start": 8,
        "Cumulative units (end": 9,
        "Headroom below Tier 1": 11,
        "Headroom below Tier 3": 13,
        "Units in no-rebate zone": 14,
        "Units in Tier-3 zone": 17,
        "Reading A": 18,
        "Reading B": 19,
        "Active Rebate": 20,
        "Warrant contra": 22,
        "Net Revenue — GAAP": 24,
        "Net Revenue — cash": 25,
        "Gross Margin — GAAP": 29,
        "Cash net billing": 37,
        "Prepayment available": 38,
        "Operating cash flow": 42,
        "NPV (pre-tax": 44,
    }
    for label, row in expected.items():
        assert _find_row(ws, label) == row, f"{label} moved off row {row}"


def test_three_tiers_reading_a_unchanged(wb3):
    ws = wb3["Model"]
    formula = str(ws.cell(_find_row(ws, "Reading A"), 2).value)
    assert formula == ("=(B15*RebateTier1Rate+B16*RebateTier2Rate"
                       "+B17*RebateTier3Rate)*ASP")
