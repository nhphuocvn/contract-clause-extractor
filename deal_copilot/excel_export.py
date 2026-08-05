"""Excel export — the "LC4 chip": a reusable, fully-live, self-documenting deal model.

``build_workbook`` turns fully-computed engine output into an openpyxl Workbook
that is BOTH a reusable deal-modeling chip AND a self-teaching document.

TWO NON-NEGOTIABLE PRINCIPLES
=============================

1. LIVENESS — zero frozen numbers in the calculation path.
   Every model INPUT is a labeled, editable cell ("pin"); every downstream number
   is a FORMULA reading those pins via workbook-scope named ranges. Change one
   input and the whole model recomputes inside Excel. NPV is a live ``=NPV()``.
   The warrant contra is a live deployment-band allocation (so reducing Demand%
   reduces the contra, exactly as the engine does). The ONLY static numeric cells
   allowed in the calc path are the labeled input pins themselves. The single
   exception is the Scenarios tab's DOWNSIDE/UPSIDE/ET columns (three scenarios
   side-by-side cannot be driven by one Demand% cell) — those are engine snapshots,
   loudly banner-labeled "change Demand% on the Model tab to explore any level live".

2. SELF-DOCUMENTATION — a reader with zero background understands it tab by tab.
   Layer 1 (visible columns): every term/assumption row shows the raw value + a
   plain-English ELI5 + the actual clause text + an example + owner + provenance.
   Layer 2 (hover notes): every calculation cell carries an openpyxl Comment
   explaining what it computes and why. Text never truncates (wrap + row height).

CIRCUIT RULE (§9.1): one labeled arithmetic step per row, read top to bottom;
formulas reference clearly-labeled named cells; many simple tabs over few clever
ones. The spreadsheet IS the engine's logic made visible.
"""

from __future__ import annotations

from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from deal_copilot.schemas import (
    AssumptionType,
    CRBMemo,
    DealAssumptions,
    DealEconomics,
    DealPackage,
    DealVersion,
    ProvenanceClass,
    RegisterEntry,
    ScenarioName,
    ViewMode,
    WarrantEconomics,
)

MODEL_VERSION = "v8.0 (Phase 7 rebuild)"
CURRENCY_NOTE = "Currency: USD, ACTUAL DOLLARS (not thousands or millions)."

# ─── Fills & fonts ───────────────────────────────────────────────────────────

_F_CONTRACT    = PatternFill("solid", fgColor="C6EFCE")  # green
_F_LIBRARY     = PatternFill("solid", fgColor="FFEB9C")  # yellow
_F_POLICY      = PatternFill("solid", fgColor="9DC3E6")  # blue
_F_JUDGMENT    = PatternFill("solid", fgColor="FFD966")  # orange
_F_PLACEHOLDER = PatternFill("solid", fgColor="FFC7CE")  # red
_F_HELPER      = PatternFill("solid", fgColor="F2F2F2")  # light gray
_F_SECTION_HDR = PatternFill("solid", fgColor="D9D9D9")
_F_TITLE       = PatternFill("solid", fgColor="1F3864")  # deep navy
_F_STATIC      = PatternFill("solid", fgColor="FFF2CC")  # amber
_F_BANNER      = PatternFill("solid", fgColor="FF9900")  # bright amber
_F_OK          = PatternFill("solid", fgColor="C6EFCE")
_F_BAD         = PatternFill("solid", fgColor="FFC7CE")

_FONT_BOLD        = Font(bold=True)
_FONT_TITLE       = Font(bold=True, color="FFFFFF", size=13)
_FONT_SUBTITLE    = Font(italic=True, color="FFFFFF")
_FONT_STATIC      = Font(color="996600")
_FONT_BANNER_BOLD = Font(bold=True, color="000000")
_FONT_ITALIC_GRAY = Font(italic=True, color="595959")
_FONT_NOTE        = Font(color="404040")

_WRAP = Alignment(wrap_text=True, vertical="top")
_WRAP_CENTER = Alignment(wrap_text=True, vertical="center")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_PROV_LABEL = {
    ProvenanceClass.CONTRACT: "Contract fact",
    ProvenanceClass.TERM_SHEET: "Term sheet",
    ProvenanceClass.LIBRARY_DEFAULT: "Library default",
    ProvenanceClass.PLACEHOLDER: "Placeholder — confirm",
}
_PROV_FILLS = {
    ProvenanceClass.CONTRACT: _F_CONTRACT,
    ProvenanceClass.TERM_SHEET: _F_CONTRACT,
    ProvenanceClass.LIBRARY_DEFAULT: _F_LIBRARY,
    ProvenanceClass.PLACEHOLDER: _F_PLACEHOLDER,
}
_ATYPE_FILLS = {
    AssumptionType.CONTRACT_FACT: _F_CONTRACT,
    AssumptionType.MARKET_DATA: _F_LIBRARY,
    AssumptionType.POLICY_NUMBER: _F_POLICY,
    AssumptionType.STRATEGIC_JUDGMENT: _F_JUDGMENT,
}

DEAL_CONFIDENTIAL = "CONFIDENTIAL — SYNTHETIC / FICTIONAL — FOR DEMO USE ONLY"

# ─── Model tab row layout (title row 1, header row 2, data from row 3) ─────────


class _ModelRows:
    """Model-tab row layout, GENERATED from the actual rebate-tier count.

    N rebate tiers produce N headroom rows and N + 1 zone rows (below tier 1,
    one band per tier, and everything above the top tier). Every row beneath the
    rebate block therefore shifts with N, so the Warrant and Acct_Sched tabs read
    their Model row numbers off this object instead of fixed constants. With the
    standard 3-tier deal the numbers are identical to the previous hardcoded
    layout, so a 3-tier workbook is byte-for-byte unchanged.

    N = 0 (a contract with no rebate at all) is legal: no headroom or zone rows
    are emitted and the two rebate readings are written as a literal zero.
    """

    TITLE = 1
    HDR = 2

    def __init__(self, n_tiers: int) -> None:
        self.n_tiers = n_tiers
        r = 3
        self.SEC_REV = r; r += 1
        self.UNITS = r; r += 1        # Demand-adjusted units = committed × Demand%
        self.ASP = r; r += 1
        self.GROSS_REV = r; r += 1
        self.SEC_DEPLOY = r; r += 1
        self.CUM_START = r; r += 1    # cumulative deployed units, start of quarter
        self.CUM_END = r; r += 1      # cumulative deployed units, end of quarter
        self.SEC_REBATE = r; r += 1
        self.HEAD = [r + i for i in range(n_tiers)]; r += n_tiers
        n_zones = n_tiers + 1 if n_tiers else 0
        self.ZONE = [r + i for i in range(n_zones)]; r += n_zones
        self.REBATE_A = r; r += 1
        self.REBATE_B = r; r += 1
        self.ACT_REBATE = r; r += 1
        self.SEC_NET = r; r += 1
        self.WARRANT_CTR = r; r += 1
        self.ADHOC = r; r += 1
        self.NET_GAAP = r; r += 1     # GAAP net revenue (carries warrant contra)
        self.NET_CASH = r; r += 1     # cash net revenue (excludes non-cash warrant)
        self.SEC_COST = r; r += 1
        self.UNIT_COGS = r; r += 1
        self.COGS = r; r += 1
        self.GM_GAAP = r; r += 1
        self.GMPCT_GAAP = r; r += 1
        self.GM_CASH = r; r += 1
        self.GMPCT_CASH = r; r += 1
        self.OPEX_PCT = r; r += 1
        self.ALLOC_OPEX = r; r += 1
        self.CONTRIB = r; r += 1
        self.SEC_CASH = r; r += 1
        self.NET_BILL = r; r += 1     # cash net billing = gross − active rebate
        self.PREPAY_AVL = r; r += 1   # prepayment available, start of quarter
        self.DRAWDOWN = r; r += 1
        self.PREPAY_END = r; r += 1
        self.COLLECTED = r; r += 1    # cash collected = net billing − drawdown
        self.OPCASH = r; r += 1       # operating cash = collected − COGS − opex
        self.WACC_Q = r; r += 1       # quarterly WACC
        self.NPV = r; r += 1          # live =NPV()


_Q1_COL = 2   # Q1 = column B

# Warrant tab pre-committed layout (Model references contra before Warrant built)
_W_VAL_COL          = 2
_W_TOTAL_EFV_ROW    = 9
_W_TR_ROW_START     = 4    # tranche table rows 4..7
_W_CONTRA_ROW       = 19   # horizontal contra-per-quarter row (cols B..M)


# ─── Generic helpers ───────────────────────────────────────────────────────────

def _cl(col: int) -> str:
    return get_column_letter(col)


def _note(cell, text: str) -> None:
    """Attach a hover comment (Layer-2 documentation) to a calculation cell."""
    c = Comment(text, "Deal Copilot")
    c.width = 360
    c.height = 200
    cell.comment = c


def _title_block(ws, n_cols: int, deal_name: str, subtitle: str, as_of: str) -> None:
    """Row-1 title block present on every tab."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1,
                value=f"{deal_name}   |   {subtitle}   |   {as_of}   |   {DEAL_CONFIDENTIAL}")
    c.fill = _F_TITLE
    c.font = _FONT_TITLE
    c.alignment = _WRAP_CENTER
    ws.row_dimensions[1].height = 30


def _section_hdr(ws, row: int, text: str, n_cols: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _F_SECTION_HDR
    c.font = _FONT_BOLD


def _banner(ws, row: int, n_cols: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _F_BANNER
    c.font = _FONT_BANNER_BOLD
    c.alignment = _WRAP_CENTER
    ws.row_dimensions[row].height = 42


def _static(cell, value: Any) -> None:
    cell.value = value
    cell.fill = _F_STATIC
    cell.font = _FONT_STATIC


def _wrapped(ws, row: int, col: int, text: str, height: int | None = None) -> None:
    # This writer only ever carries prose (notes, clause text, explanations) — never a
    # formula. A leading "=" would make Excel try to evaluate the sentence as a formula
    # (e.g. "= start + units…" → #NAME?), so strip it and present the text verbatim.
    if isinstance(text, str) and text.startswith("="):
        text = text[1:].lstrip()
    c = ws.cell(row=row, column=col, value=text)
    c.alignment = _WRAP
    c.font = _FONT_NOTE
    if height:
        ws.row_dimensions[row].height = height


# ═══════════════════════════════════════════════════════════════════════════════
# Cover tab
# ═══════════════════════════════════════════════════════════════════════════════

_LEGEND = [
    ("Green",  "Contract fact — editable input sourced from the contract",        _F_CONTRACT),
    ("Yellow", "Library default — editable; review before signing",               _F_LIBRARY),
    ("Blue",   "Policy number — confirm with Treasury / Tax",                      _F_POLICY),
    ("Orange", "Strategic judgment — confirm with the deal team",                  _F_JUDGMENT),
    ("Red",    "Placeholder — must be confirmed before use",                       _F_PLACEHOLDER),
    ("White",  "LIVE FORMULA — recalculates automatically when inputs change",     None),
    ("Amber",  "STATIC SNAPSHOT — does NOT recalculate; only on the Scenarios tab", _F_STATIC),
]

_READING_GUIDE = [
    ("Cover", "You are here. The deal in one paragraph, plus how to read the rest."),
    ("Assumptions", "Every input you can type. Change a value here and the whole model recomputes."),
    ("Warrant_Assump", "The judgment inputs for the warrant (stock price, vest odds). Strategic estimates."),
    ("Drivers", "Each contract term, the document and section it came from, and the verbatim clause text."),
    ("Model", "The quarterly P&L as a visible circuit — one arithmetic step per row, top to bottom."),
    ("Warrant", "How the free stock (warrant) is valued and turned into contra-revenue."),
    ("Scenarios", "BASE (live) vs DOWNSIDE / UPSIDE / EARLY-TERMINATION (engine snapshots)."),
    ("Acct_Sched", "Hand-off to Accounting: rebate accrual, prepayment, and a monthly receivables roll."),
    ("Variance", "Driver-by-driver walk between two saved versions (when they exist)."),
    ("Assumption_Reg", "Every input with its type and the OWNER who must confirm it."),
    ("CRB_Summary", "The one-page approval summary: verdict, approvers, risks, recommendation."),
    ("Changelog", "Audit trail of edits (when they exist)."),
    ("Analysis — Finance Manager", "A finance manager's written read of the deal."),
    ("Analysis — Plain English", "The same story with zero jargon — start here if you're new."),
]


def _write_cover(ws, pkg: DealPackage, as_of: str) -> None:
    ws.title = "Cover"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 95
    _title_block(ws, 2, pkg.deal_name, "Deal Economics Model", as_of)

    r = 3
    ws.cell(row=r, column=1, value="MODEL HEADER STAMP").font = _FONT_BOLD
    ws.cell(row=r, column=1).fill = _F_SECTION_HDR
    ws.cell(row=r, column=2).fill = _F_SECTION_HDR
    r += 1
    stamp = [
        ("Model version", MODEL_VERSION),
        ("Counterparty", pkg.counterparty or ""),
        ("Contract", "GPU Cloud Product Purchase Agreement + Warrant Agreement"),
        ("Contract effective date", "February 1, 2026"),
        ("Model as-of date", as_of),
        ("Currency / scale", CURRENCY_NOTE),
        ("NPV basis", "PRE-TAX operating cash. The tax-rate input is SHOWN for reference but is "
                      "NOT applied to the cash flows. Do not read the tax cell as 'used'."),
        ("Status", "SYNTHETIC / FICTIONAL — no real customer data; illustrative only."),
    ]
    for label, val in stamp:
        ws.cell(row=r, column=1, value=label).font = _FONT_BOLD
        _wrapped(ws, r, 2, val)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="THE DEAL IN ONE PARAGRAPH").font = _FONT_BOLD
    ws.cell(row=r, column=1).fill = _F_SECTION_HDR
    ws.cell(row=r, column=2).fill = _F_SECTION_HDR
    r += 1
    _wrapped(ws, r, 1,
             "AMD agrees to sell Meta 150,000 AI accelerator chips over three years at a "
             "$25,000 sticker price (about $3.75B of product). Meta gets volume rebates, a "
             "$500M prepayment credit, and net-90 payment terms. The twist: AMD ALSO hands "
             "Meta a warrant — the right to buy 12,000,000 AMD shares for a penny each — that "
             "vests as Meta deploys the chips. At AMD's ~$470 share price that free stock is "
             "worth about $3.4B, which is LARGER than the product's gross margin. So in cash "
             "terms the deal earns a healthy ~37% margin, but under GAAP accounting the warrant "
             "is subtracted as contra-revenue and the GAAP margin is negative. AMD is paying "
             "for this deal largely in equity, not cash.", height=150)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 2

    ws.cell(row=r, column=1, value="HOW TO READ THIS WORKBOOK (tab by tab)").font = _FONT_BOLD
    ws.cell(row=r, column=1).fill = _F_SECTION_HDR
    ws.cell(row=r, column=2).fill = _F_SECTION_HDR
    r += 1
    for tab, desc in _READING_GUIDE:
        ws.cell(row=r, column=1, value=tab).font = _FONT_BOLD
        _wrapped(ws, r, 2, desc)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="CELL COLOUR LEGEND").font = _FONT_BOLD
    ws.cell(row=r, column=1).fill = _F_SECTION_HDR
    ws.cell(row=r, column=2).fill = _F_SECTION_HDR
    r += 1
    for swatch, desc, fill in _LEGEND:
        c = ws.cell(row=r, column=1, value=swatch)
        if fill:
            c.fill = fill
        c.font = _FONT_BOLD
        _wrapped(ws, r, 2, desc)
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# Assumptions tab — every typeable input, two documentation layers
# ═══════════════════════════════════════════════════════════════════════════════

# ELI5 explanations keyed by a short tag; authored plain-English (Layer 1).
_ELI5 = {
    "ASP": "Sticker price for one GPU chip before any discounts — the price tag before coupons.",
    "Demand": "How much of the committed volume actually ships, as a %. 100 = full plan, "
              "75 = a soft-demand downside, 110 = an upside. Multiplies every quarter's units.",
    "PaymentTermsDays": "How many days Meta has to pay an invoice. Net-90 = three months — the "
                        "longer this is, the more cash AMD has tied up waiting to be paid.",
    "PrepaymentUSD": "Cash Meta pays AMD up front ($500M). It's drawn down a little against each "
                     "invoice until used up — like a gift card AMD applies to each bill.",
    "ToPFloor": "Take-or-pay floor: Meta must pay for at least this share (80%) of each year's "
                "planned units even if it doesn't take them. A revenue safety net for AMD.",
    "UnitCOGS": "What it costs AMD to build one chip (parts + manufacturing). Sticker minus this "
                "is the raw product profit per chip.",
    "OpExPct": "Operating expenses (sales, support, overhead) charged to this deal, as a % of "
               "net revenue. A rough allocation, not a line-item budget.",
    "WACC": "The discount rate — how much future dollars are worth less than today's dollars. "
            "Used to compute NPV. Higher WACC = future cash worth less now.",
    "TaxRate": "Corporate tax rate. SHOWN for reference only — this model's NPV is PRE-TAX and "
               "does NOT apply tax to the cash flows.",
    "DPODays": "How long AMD takes to pay ITS suppliers. Longer = AMD holds onto cash longer.",
    "InventoryLeadMonths": "How many months ahead AMD must buy/build chip inventory before "
                           "shipping. Ties up cash early.",
    "RebateTierThreshold": "Cumulative units Meta must buy to unlock this rebate tier.",
    "RebateTierRate": "The discount % off sticker once this tier's volume is reached.",
    "RebateToggle": "Which reading of the ambiguous rebate clause to use. A = prospective "
                    "(only units above the threshold get the higher rate). B = retroactive "
                    "(the whole year's units get the higher rate). B is the default (ASC 606 "
                    "treats a retrospective volume rebate as variable consideration). The gap "
                    "between the two is ~$41M — a Legal + Revenue Accounting question.",
    "WarrantStockPrice": "Assumed AMD share price used to value the warrant. The warrant lets "
                         "Meta buy shares for a penny, so each share is worth ~ (price − $0.01) "
                         "of free value. Higher price = more expensive warrant for AMD.",
    "VestProb": "The odds (%) this tranche of the warrant actually vests — i.e. that Meta hits "
                "the deployment milestone AND AMD's stock clears the price hurdle. A judgment call.",
    "Prob": "Your estimated probability of this scenario happening. The four must sum to 100%.",
}

_REBATE_AMBIGUITY_NOTE = (
    "AMBIGUOUS CLAUSE (§5). The contract says the higher tier applies 'based on volumes "
    "purchased during such Year' but doesn't say whether crossing a tier mid-year reprices "
    "earlier units. Reading A (prospective) = $142.5M total rebate; Reading B (retroactive) "
    "= $183.5M. The ~$41M gap is a Legal + Revenue Accounting question. Default = B."
)


def _input_row(ws, row: int, label: str, value: Any, unit: str, eli5: str,
               clause: str, example: str, owner: str, basis: ProvenanceClass,
               *, fmt: str | None = None, name: str | None = None,
               cell_map: dict | None = None, dropdown: list[str] | None = None,
               note: str = "") -> None:
    """One Assumptions row: Label | Value | Unit | ELI5 | Clause | Example | Owner | Provenance."""
    ws.cell(row=row, column=1, value=label).font = _FONT_BOLD
    c = ws.cell(row=row, column=2, value=value)
    c.protection = Protection(locked=False)        # inputs are editable
    fill = _PROV_FILLS.get(basis, _F_LIBRARY)
    c.fill = fill
    if fmt:
        c.number_format = fmt
    if note:
        _note(c, note)
    ws.cell(row=row, column=3, value=unit)
    _wrapped(ws, row, 4, eli5)
    _wrapped(ws, row, 5, clause)
    _wrapped(ws, row, 6, example)
    ws.cell(row=row, column=7, value=owner).alignment = _WRAP
    pc = ws.cell(row=row, column=8, value=_PROV_LABEL.get(basis, ""))
    pc.fill = fill
    # generous row height so wrapped clause text is never clipped
    longest = max(len(eli5), len(clause), len(example))
    ws.row_dimensions[row].height = max(30, min(150, 14 * (longest // 38 + 1)))
    if dropdown:
        dv = DataValidation(type="list", formula1='"' + ",".join(dropdown) + '"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(c)
    if name and cell_map is not None:
        cell_map[name] = f"Assumptions!${_cl(2)}${row}"


def _clause_for(terms, term_type) -> str:
    for t in terms:
        if t.term_type == term_type:
            head = t.raw_text.strip().replace("\n\n", " ")
            return f"[{t.source_document} {t.source_section}] {head}"
    return ""


def _write_assumptions(ws, assumptions: DealAssumptions, inp, pkg: DealPackage,
                       as_of: str, cell_map: dict) -> None:
    from deal_copilot.schemas import TermType
    ws.title = "Assumptions"
    widths = {"A": 34, "B": 16, "C": 10, "D": 46, "E": 60, "F": 30, "G": 26, "H": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    _title_block(ws, 8, pkg.deal_name, "Assumptions — every typeable input", as_of)

    hdr = ["Input", "Value", "Unit", "Plain-English explanation (ELI5)",
           "Clause text (source)", "Example", "Owner (who confirms)", "Provenance"]
    for col, txt in enumerate(hdr, 1):
        c = ws.cell(row=2, column=col, value=txt)
        c.font = _FONT_BOLD
        c.fill = _F_SECTION_HDR
    ws.freeze_panes = "A3"

    terms = pkg.terms
    r = 3
    _section_hdr(ws, r, "SECTION 1 — DEAL STRUCTURE (contract facts)", 8); r += 1
    _input_row(ws, r, "Base ASP", inp.base_asp, "$/unit", _ELI5["ASP"],
               _clause_for(terms, TermType.PRICING), "$25,000", "contract §4 / Sales",
               ProvenanceClass.CONTRACT, fmt="#,##0", name="ASP", cell_map=cell_map,
               note="Base average selling price per GPU. Sticker before rebates and the "
                    "warrant. Drives Gross Revenue on the Model tab."); r += 1
    _input_row(ws, r, "Demand %", 1.0, "%", _ELI5["Demand"], "(modeling control — not a contract term)",
               "100% full / 75% downside / 110% upside", "deal team",
               ProvenanceClass.PLACEHOLDER, fmt="0%", name="Demand", cell_map=cell_map,
               note="THE master volume dial. Model Units = committed schedule × this %. "
                    "Set 75% to see the downside live; the warrant contra shrinks with it."); r += 1
    _input_row(ws, r, "Payment Terms (DSO)", inp.dso_days, "days", _ELI5["PaymentTermsDays"],
               _clause_for(terms, TermType.PAYMENT_TERMS), "90 (net-90)", "contract §8",
               ProvenanceClass.CONTRACT, fmt="0", name="PaymentTermsDays", cell_map=cell_map); r += 1
    _input_row(ws, r, "Prepayment", inp.prepayment_usd, "USD", _ELI5["PrepaymentUSD"],
               _clause_for(terms, TermType.PREPAYMENT), "$500,000,000", "contract §7",
               ProvenanceClass.CONTRACT, fmt="#,##0", name="PrepaymentUSD", cell_map=cell_map); r += 1
    _input_row(ws, r, "Take-or-Pay Floor", inp.take_or_pay_floor_pct, "%", _ELI5["ToPFloor"],
               _clause_for(terms, TermType.TAKE_OR_PAY), "80%", "contract §6",
               ProvenanceClass.CONTRACT, fmt="0%", name="ToPFloor", cell_map=cell_map); r += 1

    r += 1
    _section_hdr(ws, r, "SECTION 2 — UNIT RAMP SCHEDULE (committed volume, contract §3)", 8); r += 1
    vol_clause = _clause_for(terms, TermType.VOLUME_COMMITMENT)
    for q, units in enumerate(inp.committed_quarterly):
        _input_row(ws, r, f"Q{q + 1} Committed Units", int(units), "units",
                   "Chips AMD commits to ship in this quarter (before the Demand% dial).",
                   vol_clause if q == 0 else "(see Q1 — same §3 schedule)",
                   "Q1 = 7,000", "contract §3 / Operations", ProvenanceClass.CONTRACT,
                   fmt="#,##0", name=f"UnitQ{q + 1}", cell_map=cell_map); r += 1

    r += 1
    _section_hdr(ws, r, "SECTION 3 — REBATE TERMS (contract facts + ambiguity judgment, §5)", 8); r += 1
    tiers = list(inp.rebate_tiers)
    reb_clause = _clause_for(terms, TermType.REBATE)
    for i, (thr, rate) in enumerate(tiers):
        _input_row(ws, r, f"Tier {i + 1} Threshold", int(thr), "cum units",
                   _ELI5["RebateTierThreshold"], reb_clause if i == 0 else "(see Tier 1 — §5)",
                   f"{int(thr):,} units", "contract §5", ProvenanceClass.CONTRACT,
                   fmt="#,##0", name=f"RebateTier{i + 1}Threshold", cell_map=cell_map); r += 1
    for i, (thr, rate) in enumerate(tiers):
        _input_row(ws, r, f"Tier {i + 1} Rate", rate, "% off ASP", _ELI5["RebateTierRate"],
                   "(see Tier 1 threshold — §5)", f"{rate:.0%} off", "contract §5",
                   ProvenanceClass.CONTRACT, fmt="0.00%",
                   name=f"RebateTier{i + 1}Rate", cell_map=cell_map); r += 1
    _input_row(ws, r, "★ Rebate Reading", "B-Retroactive", "", _ELI5["RebateToggle"],
               _REBATE_AMBIGUITY_NOTE, "B (default) vs A", "Legal + Revenue Accounting",
               ProvenanceClass.PLACEHOLDER, name="RebateToggle", cell_map=cell_map,
               dropdown=["A-Prospective", "B-Retroactive"],
               note=_REBATE_AMBIGUITY_NOTE); r += 1

    r += 1
    _section_hdr(ws, r, "SECTION 4 — COST & FINANCIAL PARAMETERS", 8); r += 1
    _input_row(ws, r, "Unit COGS", assumptions.unit_cogs_usd, "$/unit", _ELI5["UnitCOGS"],
               "(library default — confirm with cost accounting)", "$15,000",
               "cost accounting", ProvenanceClass.LIBRARY_DEFAULT, fmt="#,##0",
               name="UnitCOGS", cell_map=cell_map,
               note="Blended per-unit cost to build a chip. Biggest single sensitivity on margin."); r += 1
    _input_row(ws, r, "OpEx Allocation", assumptions.opex_allocation_pct, "%", _ELI5["OpExPct"],
               "(library default — FP&A judgment)", "12%", "FP&A",
               ProvenanceClass.LIBRARY_DEFAULT, fmt="0.00%", name="OpExPct", cell_map=cell_map); r += 1
    _input_row(ws, r, "WACC (discount rate)", assumptions.discount_rate_wacc, "%", _ELI5["WACC"],
               "(policy number — Treasury sets this)", "10%", "Treasury",
               ProvenanceClass.LIBRARY_DEFAULT, fmt="0.00%", name="WACC", cell_map=cell_map,
               note="Annual discount rate. The Model converts it to a quarterly rate and feeds "
                    "the live =NPV(). Change it and NPV moves."); r += 1
    _input_row(ws, r, "Tax Rate", assumptions.tax_rate, "%", _ELI5["TaxRate"],
               "(policy number — shown, NOT applied; NPV is pre-tax)", "21%", "Tax / Treasury",
               ProvenanceClass.LIBRARY_DEFAULT, fmt="0.00%", name="TaxRate", cell_map=cell_map,
               note="SHOWN FOR REFERENCE ONLY. This model's NPV is pre-tax — no tax is applied "
                    "to the cash flows."); r += 1
    _input_row(ws, r, "Supplier DPO", assumptions.supplier_payment_dpo_days, "days",
               "How long AMD takes to pay its own suppliers.", "(library default)", "60",
               "Treasury / Procurement", ProvenanceClass.LIBRARY_DEFAULT, fmt="0",
               name="DPODays", cell_map=cell_map); r += 1
    _input_row(ws, r, "Inventory Lead", assumptions.inventory_lead_months, "months",
               _ELI5["InventoryLeadMonths"], "(library default)", "3", "Operations",
               ProvenanceClass.LIBRARY_DEFAULT, fmt="0", name="InventoryLeadMonths",
               cell_map=cell_map); r += 1

    r += 1
    _section_hdr(ws, r, "SECTION 5 — SCENARIO PROBABILITIES (strategic judgment — must sum to 100%)", 8); r += 1
    prob_defaults = [("BASE", 0.50), ("DOWNSIDE", 0.25), ("UPSIDE", 0.15), ("ET", 0.10)]
    prob_first_row = r
    for nm, prob in prob_defaults:
        _input_row(ws, r, f"P({nm})", prob, "%", _ELI5["Prob"], "(deal-team judgment)",
                   "BASE 50%", "deal team", ProvenanceClass.PLACEHOLDER, fmt="0%",
                   name=f"Prob{nm}", cell_map=cell_map); r += 1
    # Sum-check + flag (live)
    sum_cell = f"{_cl(2)}{prob_first_row}:{_cl(2)}{r - 1}"
    ws.cell(row=r, column=1, value="Probabilities total").font = _FONT_BOLD
    tot = ws.cell(row=r, column=2, value=f"=SUM({sum_cell})")
    tot.number_format = "0%"
    _note(tot, "Sum of the four scenario probabilities. Must equal 100%.")
    flag = ws.cell(row=r, column=4,
                   value=f'=IF(ABS(SUM({sum_cell})-1)>0.001,"⚠ DOES NOT SUM TO 100% — FIX","OK — sums to 100%")')
    r += 1
    # Conditional red/green on the flag cell
    from openpyxl.formatting.rule import FormulaRule
    flag_addr = f"$D${r - 1}"
    ws.conditional_formatting.add(
        flag_addr, FormulaRule(formula=[f'ISNUMBER(SEARCH("DOES NOT",{flag_addr}))'],
                               fill=_F_BAD, font=Font(bold=True, color="9C0006")))
    ws.conditional_formatting.add(
        flag_addr, FormulaRule(formula=[f'ISNUMBER(SEARCH("OK",{flag_addr}))'],
                               fill=_F_OK, font=Font(bold=True, color="006100")))
    cell_map["ProbSumCell"] = f"Assumptions!$B${r - 1}"



# ═══════════════════════════════════════════════════════════════════════════════
# Warrant_Assump tab — judgment inputs + LIVE expected-value range
# ═══════════════════════════════════════════════════════════════════════════════

# Conservative / aggressive vest-probability SETS (the "Base" column reads the
# user's editable VestProbT cells, so editing a vest prob moves it live).
_CONS_PROB = [0.7, 0.5, 0.3, 0.1]
_AGGR_PROB = [1.0, 0.9, 0.7, 0.4]


def _write_warrant_assump(ws, assumptions: DealAssumptions, warrant: WarrantEconomics | None,
                          pkg: DealPackage, as_of: str, cell_map: dict) -> None:
    ws.title = "Warrant_Assump"
    for col, w in {"A": 40, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 16}.items():
        ws.column_dimensions[col].width = w
    _title_block(ws, 7, pkg.deal_name, "Warrant judgment inputs (strategic estimates)", as_of)
    _banner(ws, 2, 7, "STRATEGIC ESTIMATES — confirm every value here with the deal team before "
                      "signature. These are judgment calls, not contract facts.")

    price = (assumptions.warrant_measurement_price_usd or assumptions.current_stock_price_usd or 0.0)
    r = 4
    ws.cell(row=r, column=1, value="Measurement Stock Price").font = _FONT_BOLD
    c = ws.cell(row=r, column=2, value=price)
    c.fill = _F_JUDGMENT
    c.protection = Protection(locked=False)
    c.number_format = "#,##0.00"
    _note(c, _ELI5["WarrantStockPrice"])
    _wrapped(ws, r, 4, _ELI5["WarrantStockPrice"])
    cell_map["WarrantStockPrice"] = f"Warrant_Assump!$B${r}"
    r += 2

    _section_hdr(ws, r, "PER-TRANCHE VEST PROBABILITIES (your live estimate — drives the warrant cost)", 7)
    r += 1
    n_tr = len(warrant.tranche_valuations) if warrant else 4
    probs = list(assumptions.tranche_vest_probabilities) or [0.9, 0.7, 0.5, 0.3]
    milestones = ([tv.deployment_milestone_units for tv in warrant.tranche_valuations]
                  if warrant else [30000, 75000, 120000, 150000])
    vest_first = r
    for i in range(n_tr):
        ws.cell(row=r, column=1,
                value=f"Tranche {i + 1} vest probability (milestone {milestones[i]:,} units)")
        cc = ws.cell(row=r, column=2, value=probs[i] if i < len(probs) else 0.5)
        cc.fill = _F_JUDGMENT
        cc.protection = Protection(locked=False)
        cc.number_format = "0%"
        _note(cc, _ELI5["VestProb"])
        cell_map[f"VestProbT{i + 1}"] = f"Warrant_Assump!$B${r}"
        r += 1

    r += 1
    _section_hdr(ws, r, "EXPECTED-VALUE RANGE — LIVE (reads the price + the probabilities)", 7)
    r += 1
    head = ["Tranche", "Gross FV (shares×(price−$0.01))", "Cons prob", "Cons EV",
            "Base prob (your input)", "Base EV", "Aggr prob"]
    for col, h in enumerate(head, 1):
        ws.cell(row=r, column=col, value=h).font = _FONT_BOLD
    r += 1
    ev_first = r
    price_ref = cell_map["WarrantStockPrice"]
    shares_each = (warrant.tranche_valuations[0].share_count if warrant else 3_000_000)
    for i in range(n_tr):
        shares = warrant.tranche_valuations[i].share_count if warrant else shares_each
        ws.cell(row=r, column=1, value=f"Tranche {i + 1}")
        gfv = ws.cell(row=r, column=2, value=f"={shares}*MAX(0,{price_ref}-0.01)")
        gfv.number_format = "#,##0"
        # Conservative
        cp = ws.cell(row=r, column=3, value=_CONS_PROB[i]); cp.number_format = "0%"
        cp.fill = _F_JUDGMENT; cp.protection = Protection(locked=False)
        ws.cell(row=r, column=4, value=f"=B{r}*C{r}").number_format = "#,##0"
        # Base — reads the user's editable vest prob cell
        bp = ws.cell(row=r, column=5, value=f"={cell_map[f'VestProbT{i + 1}']}")
        bp.number_format = "0%"
        ws.cell(row=r, column=6, value=f"=B{r}*E{r}").number_format = "#,##0"
        # Aggressive
        ap = ws.cell(row=r, column=7, value=_AGGR_PROB[i]); ap.number_format = "0%"
        ap.fill = _F_JUDGMENT; ap.protection = Protection(locked=False)
        r += 1
    # Totals row
    ws.cell(row=r, column=1, value="TOTAL expected warrant value").font = _FONT_BOLD
    for col, lbl in ((4, "Cons"), (6, "Base")):
        tc = ws.cell(row=r, column=col, value=f"=SUM({_cl(col)}{ev_first}:{_cl(col)}{r - 1})")
        tc.number_format = "#,##0"
        tc.font = _FONT_BOLD
    # Aggressive total uses gross × aggr prob
    agg_terms = "+".join(f"B{ev_first + i}*G{ev_first + i}" for i in range(n_tr))
    ws.cell(row=r, column=7, value=f"={agg_terms}").number_format = "#,##0"
    cell_map["WAConsEV"] = f"Warrant_Assump!$D${r}"
    cell_map["WABaseEV"] = f"Warrant_Assump!$F${r}"
    _note(ws.cell(row=r, column=1),
          "All three totals are LIVE: change the stock price or any probability above and they "
          "recompute. The 'Base' column reads YOUR per-tranche vest probabilities; Conservative "
          "and Aggressive use the bracketing probability sets shown.")
    r += 2
    _wrapped(ws, r, 1,
             "Correlation caveat (§4): valuing the warrant with one spot price and INDEPENDENT "
             "per-tranche vest odds understates cost in the upside — deal success lifts AMD's "
             "stock, making the expensive later hurdles ($300/$400) more likely exactly when "
             "deployment milestones are hit. See CRB_Summary.", height=80)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)



# ═══════════════════════════════════════════════════════════════════════════════
# Drivers tab — terms → drivers with REAL document / section / clause text
# ═══════════════════════════════════════════════════════════════════════════════

def _write_drivers(ws, econ: DealEconomics, pkg: DealPackage, inp, as_of: str,
                   cell_map: dict) -> None:
    ws.title = "Drivers"
    for col, w in {"A": 26, "B": 30, "C": 22, "D": 70, "E": 55, "F": 16}.items():
        ws.column_dimensions[col].width = w
    _title_block(ws, 6, pkg.deal_name, "Drivers — terms, their source clause, and treatment", as_of)

    hdr = ["Driver (plain label)", "Source document", "Section",
           "Clause text (verbatim)", "Accounting / modelling note", "Value / total"]
    for col, h in enumerate(hdr, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = _FONT_BOLD
        c.fill = _F_SECTION_HDR
    ws.freeze_panes = "A3"

    term_by_id = {t.term_id: t for t in pkg.terms}
    _DRIVER_LABEL = {
        "VOLUME_RAMP": "Volume ramp (units/quarter)",
        "PRICE_ASP": "Base ASP",
        "REBATE": "Volume rebate",
        "TAKE_OR_PAY": "Take-or-pay floor",
        "PREPAYMENT": "Customer prepayment",
        "PAYMENT_TERMS": "Payment terms (DSO)",
        "WARRANT_CONTRA_REVENUE": "Warrant contra-revenue",
        "MFN": "Most-favoured-nation price protection",
        "LIABILITY_CAP": "Liability cap",
    }

    r = 3
    for d in econ.drivers:
        term = term_by_id.get(d.source_term_id)
        label = _DRIVER_LABEL.get(d.driver_type.value, d.driver_type.value.replace("_", " ").title())
        ws.cell(row=r, column=1, value=label).font = _FONT_BOLD
        ws.cell(row=r, column=2, value=(term.source_document if term else "")).alignment = _WRAP
        ws.cell(row=r, column=3, value=(term.source_section if term else "")).alignment = _WRAP
        clause = term.raw_text.strip().replace("\n\n", "\n\n") if term else ""
        _wrapped(ws, r, 4, clause)
        _wrapped(ws, r, 5, d.accounting_treatment_note or "")
        if d.value is not None:
            ws.cell(row=r, column=6, value=d.value).number_format = "#,##0"
        elif d.schedule:
            ws.cell(row=r, column=6, value=f"{len(d.schedule)}-quarter schedule")
        # height to fit the verbatim clause
        ln = max(len(clause), len(d.accounting_treatment_note or ""))
        ws.row_dimensions[r].height = max(40, min(220, 13 * (ln // 60 + 1)))
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# Model tab — the live circuit
# ═══════════════════════════════════════════════════════════════════════════════

def _write_model(ws, inp, pkg: DealPackage, as_of: str, cell_map: dict,
                 rows: _ModelRows) -> None:
    ws.title = "Model"
    n_q = len(inp.committed_quarterly)
    n_tiers = rows.n_tiers
    qpy = inp.qpy
    first_q = _Q1_COL
    last_q = first_q + n_q - 1
    total_col = last_q + 1
    note_col = total_col + 1
    n_cols = note_col

    ws.column_dimensions["A"].width = 40
    for q in range(n_q):
        ws.column_dimensions[_cl(first_q + q)].width = 12
    ws.column_dimensions[_cl(total_col)].width = 15
    ws.column_dimensions[_cl(note_col)].width = 50

    _title_block(ws, n_cols, pkg.deal_name,
                 "Model — quarterly P&L circuit (USD, actual dollars; NPV pre-tax)", as_of)

    # Header row
    ws.cell(row=rows.HDR, column=1, value="Engine step / row label").font = _FONT_BOLD
    for q in range(n_q):
        ws.cell(row=rows.HDR, column=first_q + q, value=f"Q{q + 1}").font = _FONT_BOLD
    ws.cell(row=rows.HDR, column=total_col, value="TOTAL").font = _FONT_BOLD
    ws.cell(row=rows.HDR, column=note_col, value="Note / source").font = _FONT_BOLD
    ws.freeze_panes = "B3"

    qcols = [_cl(first_q + q) for q in range(n_q)]
    tcl = _cl(total_col)

    def sum_row(row: int) -> str:
        return f"=SUM(${qcols[0]}${row}:${qcols[-1]}${row})"

    def write_row(row, label, formulas, total, note, fmt="#,##0",
                  fill=None, comment=None, unlock=False):
        ws.cell(row=row, column=1, value=label)
        for q, f in enumerate(formulas):
            c = ws.cell(row=row, column=first_q + q, value=f)
            c.number_format = fmt
            if fill:
                c.fill = fill
            if unlock:
                c.protection = Protection(locked=False)
        if total is not None:
            tc = ws.cell(row=row, column=total_col, value=total)
            tc.number_format = fmt
            if fill:
                tc.fill = fill
        if note:
            _wrapped(ws, row, note_col, note)
        if comment:
            _note(ws.cell(row=row, column=first_q), comment)

    # ── UNITS & REVENUE ──
    _section_hdr(ws, rows.SEC_REV, "— UNITS & REVENUE —", n_cols)
    write_row(rows.UNITS, "Units shipped = committed × Demand%",
              [f"=UnitQ{q + 1}*Demand" for q in range(n_q)], sum_row(rows.UNITS),
              "Committed quarterly schedule scaled by the Demand% dial (Assumptions).",
              comment="Units = committed schedule (UnitQn) × Demand%. The single master volume "
                      "control: set Demand%=75% on Assumptions and every downstream number "
                      "(revenue, rebate, COGS, warrant contra, margin, NPV) recomputes.")
    cell_map["TotalUnits"] = f"Model!${tcl}${rows.UNITS}"
    _note(ws.cell(row=rows.UNITS, column=total_col),
          "TOTAL UNITS — computed once here and reused everywhere (effective ASP, per-unit "
          "figures) via the named range 'TotalUnits'. No inline SUM is repeated elsewhere.")

    write_row(rows.ASP, "ASP ($/unit)", ["=ASP"] * n_q, "=ASP",
              "Base sticker price, constant across quarters (Assumptions §1).",
              comment="Reads the ASP named range. Change ASP on Assumptions and every quarter's "
                      "gross revenue updates.")
    write_row(rows.GROSS_REV, "Gross Revenue = Units × ASP",
              [f"={qcols[q]}{rows.UNITS}*{qcols[q]}{rows.ASP}" for q in range(n_q)],
              sum_row(rows.GROSS_REV), "Units shipped × ASP.",
              comment="Units (row above) × ASP. The top of the revenue waterfall.")
    cell_map["ModelGrossRevTotal"] = f"Model!${tcl}${rows.GROSS_REV}"

    # ── DEPLOYMENT (cumulative) ──
    _section_hdr(ws, rows.SEC_DEPLOY, "— CUMULATIVE DEPLOYMENT (drives rebate tiers & warrant vesting) —", n_cols)
    cum_start = [0] + [f"=${qcols[q - 1]}{rows.CUM_END}" for q in range(1, n_q)]
    write_row(rows.CUM_START, "Cumulative units (start of quarter)", cum_start,
              f"=${qcols[-1]}{rows.CUM_END}", "Running total deployed through the prior quarter.",
              fill=_F_HELPER,
              comment="Cumulative deployed units at the start of the quarter = end of the prior "
                      "quarter. Used both for rebate tier zones and for warrant milestone vesting.")
    write_row(rows.CUM_END, "Cumulative units (end of quarter)",
              [f"=${qcols[q]}{rows.CUM_START}+${qcols[q]}{rows.UNITS}" for q in range(n_q)],
              f"=${qcols[-1]}{rows.CUM_END}", "start + units shipped this quarter.",
              fill=_F_HELPER)

    # ── REBATE ──
    _section_hdr(ws, rows.SEC_REBATE, "— REBATE (two readings of the ambiguous §5 clause) —", n_cols)
    # One headroom row per tier: how many more units until this tier unlocks.
    for i, hrow in enumerate(rows.HEAD):
        thr_name = f"RebateTier{i + 1}Threshold"
        write_row(hrow, f"Headroom below Tier {i + 1}",
                  [f"=MAX(0,{thr_name}-${qcols[q]}{rows.CUM_START})" for q in range(n_q)],
                  None, f"MAX(0, {thr_name} − cumulative start). MAX/MIN arithmetic, no nested IFs.",
                  fill=_F_HELPER)

    # N + 1 zone rows: below tier 1, one band per tier, and above the top tier.
    # Zone[0] earns no rebate; Zone[i] (i ≥ 1) earns RebateTier{i}Rate.
    def _zone_formula(i: int, q: int) -> str:
        units = f"${qcols[q]}{rows.UNITS}"
        if i == 0:
            return f"=MIN({units},${qcols[q]}{rows.HEAD[0]})"
        if i == n_tiers:
            return f"={units}-MIN({units},${qcols[q]}{rows.HEAD[-1]})"
        return (f"=MIN({units},${qcols[q]}{rows.HEAD[i]})"
                f"-MIN({units},${qcols[q]}{rows.HEAD[i - 1]})")

    for i, zrow in enumerate(rows.ZONE):
        label = "Units in no-rebate zone" if i == 0 else f"Units in Tier-{i} zone"
        write_row(zrow, label, [_zone_formula(i, q) for q in range(n_q)], None,
                  "How many of this quarter's units fall in each rebate band (MAX/MIN arithmetic).",
                  fill=_F_HELPER)

    if n_tiers:
        reading_a = [
            "=(" + "+".join(f"{qcols[q]}{rows.ZONE[i]}*RebateTier{i}Rate"
                            for i in range(1, n_tiers + 1)) + ")*ASP"
            for q in range(n_q)
        ]
    else:
        reading_a = ["=0"] * n_q
    write_row(rows.REBATE_A, "Rebate — Reading A (prospective / marginal)",
              reading_a,
              sum_row(rows.REBATE_A), "Each unit earns the marginal rate of its cumulative band.",
              comment="Reading A (prospective): only units above a threshold get the higher rate. "
                      "Each tier's zone row is multiplied by that tier's rate, then by ASP.")
    rebate_b = []
    for q in range(n_q):
        if not n_tiers:
            rebate_b.append("=0")
            continue
        yi = q // qpy
        ye = min((yi + 1) * qpy, n_q) - 1
        rebate_b.append(f"=VLOOKUP(${qcols[ye]}{rows.CUM_END},TierTable,2,1)*${qcols[q]}{rows.UNITS}*ASP")
    write_row(rows.REBATE_B, "Rebate — Reading B (retroactive within year)",
              rebate_b, sum_row(rows.REBATE_B),
              "Year-end cumulative sets one blended rate applied to the whole year (VLOOKUP).",
              comment="Reading B (retroactive): the year-end cumulative volume sets the rate for "
                      "ALL of that year's units. Totals ~$183.5M — about $41M more than Reading A. "
                      "Default. The gap is a Legal + Revenue Accounting question.")
    write_row(rows.ACT_REBATE, "★ Active Rebate (toggle-selected)",
              [f'=IF(RebateToggle="B-Retroactive",${qcols[q]}{rows.REBATE_B},${qcols[q]}{rows.REBATE_A})'
               for q in range(n_q)], sum_row(rows.ACT_REBATE),
              "Single IF on the Rebate Reading toggle. Flip the toggle to switch readings live.",
              comment="One IF on RebateToggle picks Reading A or B. Change the toggle on "
                      "Assumptions and this row (and net revenue, margin, NPV) all move.")
    cell_map["ModelActiveRebateTotal"] = f"Model!${tcl}${rows.ACT_REBATE}"

    # ── NET REVENUE ──
    _section_hdr(ws, rows.SEC_NET, "— NET REVENUE (GAAP carries the warrant; cash view excludes it) —", n_cols)
    write_row(rows.WARRANT_CTR, "Warrant contra-revenue (non-cash)",
              [f"={cell_map[f'ContraSchedQ{q + 1}']}" for q in range(n_q)],
              sum_row(rows.WARRANT_CTR),
              "Free stock given to the customer, allocated as units cross warrant milestones "
              "(reads the Warrant tab). Non-cash. Scales with Demand%.",
              comment="ASC 606: equity to a customer is consideration payable to a customer → a "
                      "reduction of transaction price (contra-revenue), measured under ASC 718. "
                      "Reads the Warrant tab's deployment-band allocation, so fewer deployed "
                      "units = less contra. This is why the −$2B mis-scaling cannot happen here.")
    cell_map["ModelWarrantContraTotal"] = f"Model!${tcl}${rows.WARRANT_CTR}"
    write_row(rows.ADHOC, "Ad-hoc adjustment (analyst-editable)",
              [int(inp.adhoc_schedule[q]) if q < len(inp.adhoc_schedule) and inp.adhoc_schedule[q] else 0
               for q in range(n_q)], sum_row(rows.ADHOC),
              "Manual line item (side-letter credit, etc.). Positive = adds to net revenue.",
              unlock=True,
              comment="The only editable numeric cells on the Model tab. Analyst override; flows "
                      "through net revenue and margin like any driver.")
    write_row(rows.NET_GAAP, "Net Revenue — GAAP",
              [f"=${qcols[q]}{rows.GROSS_REV}-${qcols[q]}{rows.ACT_REBATE}-${qcols[q]}{rows.WARRANT_CTR}"
               f"+${qcols[q]}{rows.ADHOC}" for q in range(n_q)], sum_row(rows.NET_GAAP),
              "Gross − active rebate − warrant contra + ad-hoc. Carries the non-cash warrant.",
              comment="GAAP net revenue. Because the $3.4B warrant contra exceeds product margin, "
                      "this is far below the cash view — the headline of the deal.")
    cell_map["ModelNetGaapTotal"] = f"Model!${tcl}${rows.NET_GAAP}"
    write_row(rows.NET_CASH, "Net Revenue — cash / commercial",
              [f"=${qcols[q]}{rows.GROSS_REV}-${qcols[q]}{rows.ACT_REBATE}+${qcols[q]}{rows.ADHOC}"
               for q in range(n_q)], sum_row(rows.NET_CASH),
              "Gross − active rebate + ad-hoc. Excludes the non-cash warrant (the 'cash-economic' view).",
              comment="Cash / commercial net revenue — the non-GAAP-style view that excludes the "
                      "non-cash warrant (analogous to how AMD's own non-GAAP excludes stock comp).")
    cell_map["ModelNetCashTotal"] = f"Model!${tcl}${rows.NET_CASH}"

    # ── COST & MARGIN ──
    _section_hdr(ws, rows.SEC_COST, "— COST & MARGIN —", n_cols)
    write_row(rows.UNIT_COGS, "Unit COGS ($/unit)", ["=UnitCOGS"] * n_q, "=UnitCOGS",
              "Cost to build one chip (Assumptions §4).")
    write_row(rows.COGS, "COGS = Units × Unit COGS",
              [f"=${qcols[q]}{rows.UNITS}*${qcols[q]}{rows.UNIT_COGS}" for q in range(n_q)],
              sum_row(rows.COGS), "Units × unit cost.")
    cell_map["ModelCOGSTotal"] = f"Model!${tcl}${rows.COGS}"
    write_row(rows.GM_GAAP, "Gross Margin — GAAP",
              [f"=${qcols[q]}{rows.NET_GAAP}-${qcols[q]}{rows.COGS}" for q in range(n_q)],
              sum_row(rows.GM_GAAP), "GAAP net revenue − COGS. Negative here (warrant-dominated).",
              comment="GAAP gross margin. Negative because the warrant contra exceeds product "
                      "margin — correct, not a bug. The deal is GAAP-dilutive.")
    cell_map["ModelGmGaapTotal"] = f"Model!${tcl}${rows.GM_GAAP}"
    write_row(rows.GMPCT_GAAP, "Gross Margin % — GAAP",
              [f"=IF(${qcols[q]}{rows.NET_GAAP}=0,0,${qcols[q]}{rows.GM_GAAP}/${qcols[q]}{rows.NET_GAAP})"
               for q in range(n_q)],
              f"=IF(${tcl}{rows.NET_GAAP}=0,0,${tcl}{rows.GM_GAAP}/${tcl}{rows.NET_GAAP})",
              "GAAP gross margin ÷ GAAP net revenue.", fmt="0.0%")
    cell_map["ModelGmPctGaapTotal"] = f"Model!${tcl}${rows.GMPCT_GAAP}"
    write_row(rows.GM_CASH, "Gross Margin — cash / commercial",
              [f"=${qcols[q]}{rows.NET_CASH}-${qcols[q]}{rows.COGS}" for q in range(n_q)],
              sum_row(rows.GM_CASH), "Cash net revenue − COGS. The healthy ~37% headline margin.",
              comment="Cash gross margin — excludes the non-cash warrant. This is the ~$1.36B / "
                      "~37.6% figure the deal team quotes.")
    cell_map["ModelGmCashTotal"] = f"Model!${tcl}${rows.GM_CASH}"
    write_row(rows.GMPCT_CASH, "Gross Margin % — cash / commercial",
              [f"=IF(${qcols[q]}{rows.NET_CASH}=0,0,${qcols[q]}{rows.GM_CASH}/${qcols[q]}{rows.NET_CASH})"
               for q in range(n_q)],
              f"=IF(${tcl}{rows.NET_CASH}=0,0,${tcl}{rows.GM_CASH}/${tcl}{rows.NET_CASH})",
              "Cash gross margin ÷ cash net revenue.", fmt="0.0%")
    cell_map["ModelGmPctCashTotal"] = f"Model!${tcl}${rows.GMPCT_CASH}"
    write_row(rows.OPEX_PCT, "OpEx allocation %", ["=OpExPct"] * n_q, "=OpExPct",
              "Allocated operating expense as a % of cash net revenue.", fmt="0.00%")
    write_row(rows.ALLOC_OPEX, "Allocated OpEx",
              [f"=${qcols[q]}{rows.NET_CASH}*${qcols[q]}{rows.OPEX_PCT}" for q in range(n_q)],
              sum_row(rows.ALLOC_OPEX), "Cash net revenue × opex %.")
    write_row(rows.CONTRIB, "Contribution margin (cash)",
              [f"=${qcols[q]}{rows.GM_CASH}-${qcols[q]}{rows.ALLOC_OPEX}" for q in range(n_q)],
              sum_row(rows.CONTRIB), "Cash gross margin − allocated opex.")

    # ── CASH FLOW & NPV (live) ──
    _section_hdr(ws, rows.SEC_CASH, "— CASH FLOW & NPV (live; pre-tax; simplified quarterly timing) —", n_cols)
    write_row(rows.NET_BILL, "Cash net billing = Gross − active rebate",
              [f"=${qcols[q]}{rows.GROSS_REV}-${qcols[q]}{rows.ACT_REBATE}" for q in range(n_q)],
              sum_row(rows.NET_BILL), "What AMD invoices the customer (excludes the non-cash warrant).")
    prepay_avail = ["=PrepaymentUSD"] + [f"=${qcols[q - 1]}{rows.PREPAY_END}" for q in range(1, n_q)]
    write_row(rows.PREPAY_AVL, "Prepayment available (start of quarter)", prepay_avail,
              None, "Remaining customer prepayment at the start of the quarter.", fill=_F_HELPER,
              comment="The $500M prepayment is drawn down 20% of each invoice until exhausted "
                      "(contract §7). This row carries the running balance.")
    write_row(rows.DRAWDOWN, "Prepayment drawdown (20% of invoice)",
              [f"=MIN(${qcols[q]}{rows.PREPAY_AVL},0.2*${qcols[q]}{rows.GROSS_REV})" for q in range(n_q)],
              sum_row(rows.DRAWDOWN), "20% of the invoice, capped at the remaining prepayment.",
              fill=_F_HELPER)
    write_row(rows.PREPAY_END, "Prepayment remaining (end of quarter)",
              [f"=${qcols[q]}{rows.PREPAY_AVL}-${qcols[q]}{rows.DRAWDOWN}" for q in range(n_q)],
              None, "available − drawdown.", fill=_F_HELPER)
    write_row(rows.COLLECTED, "Customer cash collected = billing − drawdown",
              [f"=${qcols[q]}{rows.NET_BILL}-${qcols[q]}{rows.DRAWDOWN}" for q in range(n_q)],
              sum_row(rows.COLLECTED), "Cash actually received from the customer this quarter "
              "(the rest is covered by the prepayment).")
    write_row(rows.OPCASH, "Operating cash flow = collected − COGS − OpEx",
              [f"=${qcols[q]}{rows.COLLECTED}-${qcols[q]}{rows.COGS}-${qcols[q]}{rows.ALLOC_OPEX}"
               for q in range(n_q)], sum_row(rows.OPCASH),
              "Simplified: revenue collected in-quarter (no DSO lag here — the Acct_Sched tab "
              "shows the receivables timing). The prepayment inflow is added in the NPV below.",
              comment="Operating cash per quarter. SIMPLIFIED quarterly timing: the engine's "
                      "headline NPV uses a finer monthly DSO/DPO/inventory model. This in-sheet "
                      "NPV is a live pre-tax proxy that responds to every input.")
    # Quarterly WACC + live NPV
    ws.cell(row=rows.WACC_Q, column=1, value="Quarterly WACC = (1+WACC)^(1/4) − 1")
    wq = ws.cell(row=rows.WACC_Q, column=first_q, value="=(1+WACC)^0.25-1")
    wq.number_format = "0.000%"
    cell_map["ModelWaccQ"] = f"Model!${qcols[0]}${rows.WACC_Q}"
    _note(wq, "Annual WACC converted to a quarterly discount rate for =NPV(). Live on WACC.")
    _wrapped(ws, rows.WACC_Q, note_col, "Converts the annual WACC named range to a quarterly rate.")
    ws.cell(row=rows.NPV, column=1, value="NPV (pre-tax, live)").font = _FONT_BOLD
    npv_cell = ws.cell(
        row=rows.NPV, column=first_q,
        value=f"=PrepaymentUSD+NPV(${qcols[0]}${rows.WACC_Q},${qcols[0]}${rows.OPCASH}:${qcols[-1]}${rows.OPCASH})")
    npv_cell.number_format = "#,##0"
    npv_cell.font = _FONT_BOLD
    _note(npv_cell,
          "LIVE =NPV(): prepayment received at deal start, plus the operating-cash row discounted "
          "at the quarterly WACC. Change WACC, ASP, COGS, units or Demand% and this moves. "
          "Pre-tax; simplified quarterly timing vs the engine's monthly model.")
    cell_map["ModelNPV"] = f"Model!${qcols[0]}${rows.NPV}"
    _wrapped(ws, rows.NPV, note_col, "Live NPV — referenced by the Scenarios and CRB tabs.")

    # ── TAKE-OR-PAY CHECK (annual floor; live on ToPFloor and Demand%) ──
    r = rows.NPV + 2
    _section_hdr(ws, r, "— TAKE-OR-PAY CHECK (annual 80% floor; live on Demand%) —", n_cols)
    r += 1
    for col, h in enumerate(["Year", "Committed", "Floor (ToP%)", "Taken (after Demand%)",
                             "Shortfall units", "Shortfall billing ($)"], 1):
        ws.cell(row=r, column=col, value=h).font = _FONT_BOLD
    r += 1
    n_years = (n_q + qpy - 1) // qpy
    for y in range(n_years):
        q0, q1 = y * qpy, min((y + 1) * qpy, n_q)
        committed_terms = "+".join(f"UnitQ{q + 1}" for q in range(q0, q1))
        taken_terms = "+".join(f"${qcols[q]}${rows.UNITS}" for q in range(q0, q1))
        ws.cell(row=r, column=1, value=f"Year {y + 1}")
        ws.cell(row=r, column=2, value=f"={committed_terms}").number_format = "#,##0"
        ws.cell(row=r, column=3, value=f"=ToPFloor*B{r}").number_format = "#,##0"
        ws.cell(row=r, column=4, value=f"={taken_terms}").number_format = "#,##0"
        ws.cell(row=r, column=5, value=f"=MAX(0,C{r}-D{r})").number_format = "#,##0"
        ws.cell(row=r, column=6, value=f"=E{r}*ASP").number_format = "#,##0"
        if y == 0:
            _note(ws.cell(row=r, column=3),
                  "Floor = take-or-pay % × committed units for the year. If the customer takes "
                  "fewer than this (Demand% below the floor), it still pays for the shortfall at "
                  "ASP (contract §6). Live on ToPFloor and Demand% — set Demand%=75% to see the "
                  "shortfall appear. The engine folds this into the DOWNSIDE scenario's billing.")
        r += 1



# ═══════════════════════════════════════════════════════════════════════════════
# Warrant tab — valuation + LIVE deployment-band contra allocation
# ═══════════════════════════════════════════════════════════════════════════════

def _write_warrant(ws, warrant: WarrantEconomics | None, inp, pkg: DealPackage,
                   as_of: str, cell_map: dict, rows: _ModelRows) -> None:
    ws.title = "Warrant"
    n_q = len(inp.committed_quarterly)
    for col in range(1, 14):
        ws.column_dimensions[_cl(col)].width = 13
    ws.column_dimensions["A"].width = 30
    _title_block(ws, 13, pkg.deal_name, "Warrant — valuation & deployment-band contra", as_of)

    if warrant is None:
        ws.cell(row=2, column=1,
                value="No warrant on this deal — contra-revenue is zero.").font = _FONT_ITALIC_GRAY
        for q in range(n_q):
            ws.cell(row=_W_CONTRA_ROW, column=_Q1_COL + q, value=0)
        return

    # ── Tranche valuation table (live on price + vest probs) ──
    hdr = ["Tranche", "Shares", "Strike", "Hurdle", "Milestone", "Prev milestone",
           "Band width", "Vest prob", "FV/share", "Gross FV", "Expected FV", "Per-unit EFV"]
    for col, h in enumerate(hdr, 1):
        ws.cell(row=3, column=col, value=h).font = _FONT_BOLD
    price_ref = cell_map["WarrantStockPrice"]
    prev = 0
    tr_rows = []
    for i, tv in enumerate(warrant.tranche_valuations):
        r = _W_TR_ROW_START + i
        tr_rows.append(r)
        ws.cell(row=r, column=1, value=f"Tranche {i + 1}")
        ws.cell(row=r, column=2, value=tv.share_count).number_format = "#,##0"
        ws.cell(row=r, column=3, value=tv.exercise_price_usd).number_format = "#,##0.00"
        ws.cell(row=r, column=4, value=tv.stock_price_hurdle_usd).number_format = "#,##0"
        ws.cell(row=r, column=5, value=tv.deployment_milestone_units).number_format = "#,##0"
        ws.cell(row=r, column=6, value=prev).number_format = "#,##0"
        ws.cell(row=r, column=7, value=f"=E{r}-F{r}").number_format = "#,##0"
        vp = ws.cell(row=r, column=8, value=f"={cell_map[f'VestProbT{i + 1}']}")
        vp.number_format = "0%"
        ws.cell(row=r, column=9, value=f"=MAX(0,{price_ref}-C{r})").number_format = "#,##0.00"
        ws.cell(row=r, column=10, value=f"=B{r}*I{r}").number_format = "#,##0"
        ws.cell(row=r, column=11, value=f"=J{r}*H{r}").number_format = "#,##0"
        ws.cell(row=r, column=12, value=f"=IF(G{r}=0,0,K{r}/G{r})").number_format = "#,##0.00"
        prev = tv.deployment_milestone_units
    _note(ws.cell(row=_W_TR_ROW_START, column=9),
          "Fair value per share = max(0, measurement price − $0.01 strike). The warrant is "
          "near-zero-strike, so it is essentially free stock.")
    _note(ws.cell(row=_W_TR_ROW_START, column=11),
          "Expected FV = shares × FV/share × vest probability. Live on the stock price and your "
          "per-tranche vest odds.")
    _note(ws.cell(row=_W_TR_ROW_START, column=12),
          "Per-unit EFV = this tranche's expected value spread evenly over its deployment band "
          "(prev milestone → this milestone). Used to allocate contra by units deployed.")

    # ── Total EFV (named) ──
    ws.cell(row=_W_TOTAL_EFV_ROW, column=1, value="Total Expected Fair Value").font = _FONT_BOLD
    efv = ws.cell(row=_W_TOTAL_EFV_ROW, column=_W_VAL_COL,
                  value=f"=SUM(K{tr_rows[0]}:K{tr_rows[-1]})")
    efv.number_format = "#,##0"
    efv.font = _FONT_BOLD
    cell_map["WarrantTotalEFV"] = f"Warrant!${_cl(_W_VAL_COL)}${_W_TOTAL_EFV_ROW}"
    _note(efv, "Sum of the four tranches' expected fair values. At ~$470 spot ≈ $3.38B — larger "
               "than the deal's product gross margin. This is the headline warrant cost.")

    # ── Deployment-band contra allocation matrix (tranches × quarters) ──
    qcols = [_cl(_Q1_COL + q) for q in range(n_q)]
    mrow = 12
    ws.cell(row=mrow, column=1, value="CONTRA ALLOCATION — units deployed in each tranche band").font = _FONT_BOLD
    mrow += 1
    ws.cell(row=mrow, column=1, value="(quarter →)").font = _FONT_ITALIC_GRAY
    for q in range(n_q):
        ws.cell(row=mrow, column=_Q1_COL + q, value=f"Q{q + 1}").font = _FONT_BOLD
    alloc_first = mrow + 1
    for i, tr in enumerate(tr_rows):
        rr = alloc_first + i
        ws.cell(row=rr, column=1, value=f"Tranche {i + 1} contra")
        for q in range(n_q):
            cs = f"Model!${qcols[q]}${rows.CUM_START}"
            ce = f"Model!${qcols[q]}${rows.CUM_END}"
            # overlap of (cum_start, cum_end] with (prev_milestone, milestone] × per-unit EFV
            f = (f"=MAX(0,MIN({ce},$E${tr})-MAX({cs},$F${tr}))*$L${tr}")
            ws.cell(row=rr, column=_Q1_COL + q, value=f).number_format = "#,##0"
        if i == 0:
            _note(ws.cell(row=rr, column=_Q1_COL),
                  "Units of THIS quarter's cumulative deployment that fall in this tranche's band, "
                  "× per-unit EFV. MAX/MIN overlap arithmetic — the same idea as the rebate zones. "
                  "Because it keys off cumulative DEPLOYED units (which scale with Demand%), a "
                  "downside automatically yields less contra — the engine's exact logic, live.")
    alloc_last = alloc_first + len(tr_rows) - 1

    # ── Contra per quarter (the ContraSchedQn row Model reads) ──
    ws.cell(row=_W_CONTRA_ROW, column=1, value="Warrant contra per quarter (→ Model)").font = _FONT_BOLD
    for q in range(n_q):
        col = _Q1_COL + q
        cell = ws.cell(row=_W_CONTRA_ROW, column=col,
                       value=f"=SUM({_cl(col)}{alloc_first}:{_cl(col)}{alloc_last})")
        cell.number_format = "#,##0"
        cell_map[f"ContraSchedQ{q + 1}"] = f"Warrant!${_cl(col)}${_W_CONTRA_ROW}"
    _note(ws.cell(row=_W_CONTRA_ROW, column=_Q1_COL),
          "Sum of the four tranche allocations for the quarter. This is what the Model's "
          "warrant contra-revenue row reads. Note: the Model SUBTRACTS this (it is a reduction "
          "of revenue), so it appears positive here and negative in net revenue.")

    # ── Effective ASP waterfall (live) ──
    r = _W_CONTRA_ROW + 2
    _section_hdr(ws, r, "EFFECTIVE NET ASP WATERFALL (per unit)", 4); r += 1
    waterfall = [
        ("Sticker ASP", "=ASP"),
        ("Less: avg rebate / unit", "=-ModelActiveRebateTotal/TotalUnits"),
        ("Less: warrant cost / unit", "=-WarrantTotalEFV/TotalUnits"),
        ("All-in net ASP", "=ASP-ModelActiveRebateTotal/TotalUnits-WarrantTotalEFV/TotalUnits"),
    ]
    for label, f in waterfall:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=f).number_format = "#,##0.00"
        r += 1
    _note(ws.cell(row=r - 1, column=2),
          "Sticker minus per-unit rebate minus per-unit warrant cost. All live; per-unit figures "
          "use the single TotalUnits cell from the Model tab.")

    # ── GAAP vs cash bridge (live) ──
    r += 1
    _section_hdr(ws, r, "GAAP ↔ CASH BRIDGE", 4); r += 1
    bridge = [
        ("Cash / commercial net revenue", "=ModelNetCashTotal"),
        ("Less: warrant contra (non-cash)", "=-ModelWarrantContraTotal"),
        ("GAAP net revenue", "=ModelNetGaapTotal"),
    ]
    for label, f in bridge:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=f).number_format = "#,##0"
        r += 1

    r += 1
    _wrapped(ws, r, 1,
             "⚠ CORRELATION CAVEAT (§4): the valuation uses a single spot price with INDEPENDENT "
             "per-tranche vest probabilities. In reality deployment milestones and stock-price "
             "hurdles are POSITIVELY correlated — deal success lifts AMD's stock, making the "
             "expensive later hurdles more likely just as deployment milestones are hit. So the "
             "model likely UNDERSTATES warrant cost in the upside.", height=80)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)



# ═══════════════════════════════════════════════════════════════════════════════
# Drivers helper: rebate tier table for Reading-B VLOOKUP (named TierTable)
# ═══════════════════════════════════════════════════════════════════════════════

def _write_tier_table(ws_drivers, inp, cell_map: dict) -> None:
    """Append the rebate tier table used by Model Reading-B VLOOKUP."""
    ws = ws_drivers
    r = ws.max_row + 3
    _section_hdr(ws, r, "REBATE TIER TABLE — used by VLOOKUP in Model Reading-B", 6); r += 1
    ws.cell(row=r, column=1, value="Cumulative threshold").font = _FONT_BOLD
    ws.cell(row=r, column=2, value="Rate").font = _FONT_BOLD
    r += 1
    start = r
    ws.cell(row=r, column=1, value=0)
    ws.cell(row=r, column=2, value=0.0).number_format = "0.00%"
    r += 1
    for i in range(len(inp.rebate_tiers)):
        ws.cell(row=r, column=1, value=f"=RebateTier{i + 1}Threshold")
        ws.cell(row=r, column=2, value=f"=RebateTier{i + 1}Rate").number_format = "0.00%"
        r += 1
    cell_map["TierTable"] = f"Drivers!$A${start}:$B${r - 1}"


# ═══════════════════════════════════════════════════════════════════════════════
# Scenarios tab
# ═══════════════════════════════════════════════════════════════════════════════

_SCEN_ORDER = [
    (ScenarioName.BASE, "BASE", "ProbBASE"),
    (ScenarioName.DOWNSIDE_TAKE_OR_PAY, "DOWNSIDE (take-or-pay floor)", "ProbDOWNSIDE"),
    (ScenarioName.UPSIDE_VOLUME, "UPSIDE (+15% volume)", "ProbUPSIDE"),
    (ScenarioName.EARLY_TERMINATION, "EARLY TERMINATION", "ProbET"),
]


def _write_scenarios(ws, econ: DealEconomics, pkg: DealPackage, as_of: str, cell_map: dict) -> None:
    ws.title = "Scenarios"
    for col, w in {"A": 32, "B": 18, "C": 18, "D": 16, "E": 18, "F": 16, "G": 14, "H": 16}.items():
        ws.column_dimensions[col].width = w
    _title_block(ws, 8, pkg.deal_name, "Scenarios — BASE live; others engine snapshots", as_of)

    _banner(ws, 2, 8,
            "DOWNSIDE / UPSIDE / EARLY-TERMINATION below are STATIC engine snapshots — change "
            "Demand% on the Model tab to explore any volume level LIVE. Only the BASE row is live.")

    hdr = ["Scenario", "Cash net rev", "Cash GM", "Cash GM%",
           "GAAP net rev", "GAAP GM", "Prob", "EV (prob × cash GM)"]
    for col, h in enumerate(hdr, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _FONT_BOLD
        c.fill = _F_SECTION_HDR

    gaap = {r.scenario: r for r in econ.scenarios if r.view == ViewMode.GAAP}
    cash = {r.scenario: r for r in econ.scenarios if r.view == ViewMode.CASH_COMMERCIAL}
    r = 4
    rows_for_ev = []
    for sname, label, prob_name in _SCEN_ORDER:
        ws.cell(row=r, column=1, value=label).font = _FONT_BOLD
        if sname == ScenarioName.BASE:
            ws.cell(row=r, column=2, value="=ModelNetCashTotal").number_format = "#,##0"
            ws.cell(row=r, column=3, value="=ModelGmCashTotal").number_format = "#,##0"
            ws.cell(row=r, column=4, value="=ModelGmPctCashTotal").number_format = "0.0%"
            ws.cell(row=r, column=5, value="=ModelNetGaapTotal").number_format = "#,##0"
            ws.cell(row=r, column=6, value="=ModelGmGaapTotal").number_format = "#,##0"
        else:
            cs, gs = cash.get(sname), gaap.get(sname)
            _static(ws.cell(row=r, column=2), round(cs.total_net_revenue) if cs else 0)
            _static(ws.cell(row=r, column=3), round(cs.total_gross_margin) if cs else 0)
            _static(ws.cell(row=r, column=4), cs.total_gross_margin_pct if cs else 0)
            _static(ws.cell(row=r, column=5), round(gs.total_net_revenue) if gs else 0)
            _static(ws.cell(row=r, column=6), round(gs.total_gross_margin) if gs else 0)
            for col in (2, 3, 5, 6):
                ws.cell(row=r, column=col).number_format = "#,##0"
            ws.cell(row=r, column=4).number_format = "0.0%"
        ws.cell(row=r, column=7, value=f"={prob_name}").number_format = "0%"
        ws.cell(row=r, column=8, value=f"=C{r}*G{r}").number_format = "#,##0"
        rows_for_ev.append(r)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Probability-weighted expected cash GM").font = _FONT_BOLD
    ws.cell(row=r, column=8,
            value="=" + "+".join(f"H{rr}" for rr in rows_for_ev)).number_format = "#,##0"
    r += 2

    # The loud cash→GAAP bridge panel (required framing)
    _banner(ws, r, 8, "WHY CASH AND GAAP DIFFER — READ THIS")
    r += 1
    _wrapped(ws, r, 1,
             "Cash margin ~37.6% is healthy; GAAP margin is negative because the $3.38B warrant "
             "(free stock to the customer) exceeds the $1.36B product margin and is subtracted as "
             "contra-revenue. The deal is cash-positive but GAAP-dilutive — AMD pays for it in "
             "equity, not cash. The cash view is the 'non-GAAP / cash-economic' view (it excludes "
             "the non-cash warrant, much like AMD's own non-GAAP results exclude stock-based "
             "compensation); the GAAP view carries the full warrant contra.", height=110)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    ws.cell(row=r, column=1,
            value=f"Engine snapshot generated: {as_of}.").font = _FONT_ITALIC_GRAY



# ═══════════════════════════════════════════════════════════════════════════════
# Accounting schedules tab
# ═══════════════════════════════════════════════════════════════════════════════

def _write_acct_sched(ws, inp, pkg: DealPackage, as_of: str, cell_map: dict,
                      rows: _ModelRows) -> None:
    ws.title = "Acct_Sched"
    n_q = len(inp.committed_quarterly)
    qpy = inp.qpy
    ws.column_dimensions["A"].width = 30
    for col in range(2, 8):
        ws.column_dimensions[_cl(col)].width = 16
    _title_block(ws, 6, pkg.deal_name, "Accounting schedules — accrual, prepayment, receivables", as_of)

    r = 2
    _section_hdr(ws, r, "REBATE ACCRUAL WALK (ASC 606 variable consideration)", 5); r += 1
    for col, h in enumerate(["Quarter", "Beginning", "Accrual", "Settlement", "Ending"], 1):
        ws.cell(row=r, column=col, value=h).font = _FONT_BOLD
    r += 1
    for q in range(n_q):
        qc = _cl(_Q1_COL + q)
        ws.cell(row=r, column=1, value=f"Q{q + 1}")
        ws.cell(row=r, column=2, value=(0 if q == 0 else f"=E{r - 1}")).number_format = "#,##0"
        ws.cell(row=r, column=3, value=f"=Model!${qc}${rows.ACT_REBATE}").number_format = "#,##0"
        is_year_end = ((q + 1) % qpy == 0) or (q == n_q - 1)
        if is_year_end:
            ys = r - (q % qpy)
            ws.cell(row=r, column=4, value=f"=SUM(C{ys}:C{r})").number_format = "#,##0"
        else:
            ws.cell(row=r, column=4, value=0).number_format = "#,##0"
        ws.cell(row=r, column=5, value=f"=B{r}+C{r}-D{r}").number_format = "#,##0"
        r += 1
    _wrapped(ws, r, 1, "Continuity: ending[q] = beginning[q+1]. Settlement annual-in-arrears "
                       "(Q4/Q8/Q12). Accrual reads the Model's active rebate row, so it follows "
                       "the A/B toggle live."); r += 2

    _section_hdr(ws, r, "CONTRACT LIABILITY — PREPAYMENT SCHEDULE", 4); r += 1
    for col, h in enumerate(["Quarter", "Beginning", "Drawdown", "Ending"], 1):
        ws.cell(row=r, column=col, value=h).font = _FONT_BOLD
    r += 1
    for q in range(n_q):
        qc = _cl(_Q1_COL + q)
        ws.cell(row=r, column=1, value=f"Q{q + 1}")
        ws.cell(row=r, column=2, value=f"=Model!${qc}${rows.PREPAY_AVL}").number_format = "#,##0"
        ws.cell(row=r, column=3, value=f"=Model!${qc}${rows.DRAWDOWN}").number_format = "#,##0"
        ws.cell(row=r, column=4, value=f"=Model!${qc}${rows.PREPAY_END}").number_format = "#,##0"
        r += 1
    _wrapped(ws, r, 1, "Mirrors the Model's live prepayment roll-forward (20% of each invoice "
                       "until exhausted)."); r += 2

    # ── Peak receivables — live monthly AR roll-forward ──
    _section_hdr(ws, r, "PEAK RECEIVABLES — monthly AR roll-forward (live)", 8); r += 1
    dso_months = max(1, round(inp.dso_days / 30.4375))
    _wrapped(ws, r,
             1, f"Monthly accounts-receivable balance. Billing is spread evenly across each "
                f"quarter's 3 months and collected {dso_months} month(s) later (net-{inp.dso_days}). "
                f"AR = cumulative billed − cumulative collected; peak AR is the worst month. "
                f"Billing is fully live (scales with units, ASP, Demand%); the {dso_months}-month "
                f"collection LAG is fixed at export — change Payment Terms and rerun to rebuild it.",
             height=70)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    n_m = n_q * 3
    hdr_row = r
    ws.cell(row=r, column=1, value="Month").font = _FONT_BOLD
    bill_row = r + 1
    coll_row = r + 2
    ar_row = r + 3
    ws.cell(row=bill_row, column=1, value="Monthly billing")
    ws.cell(row=coll_row, column=1, value="Monthly collection")
    ws.cell(row=ar_row, column=1, value="AR balance (cum billed − collected)")
    # months laid out across columns B...
    mcols = [_cl(2 + m) for m in range(n_m)]
    for m in range(n_m):
        q = m // 3
        qc = _cl(_Q1_COL + q)
        col = 2 + m
        ws.cell(row=hdr_row, column=col, value=f"M{m + 1}").font = _FONT_BOLD
        # billing = quarter net billing / 3
        ws.cell(row=bill_row, column=col,
                value=f"=Model!${qc}${rows.NET_BILL}/3").number_format = "#,##0"
        # collection = billing dso_months earlier (0 before that)
        if m - dso_months >= 0:
            ws.cell(row=coll_row, column=col,
                    value=f"={mcols[m - dso_months]}{bill_row}").number_format = "#,##0"
        else:
            ws.cell(row=coll_row, column=col, value=0).number_format = "#,##0"
        # AR balance = prior AR + billing − collection
        if m == 0:
            ws.cell(row=ar_row, column=col,
                    value=f"={mcols[m]}{bill_row}-{mcols[m]}{coll_row}").number_format = "#,##0"
        else:
            ws.cell(row=ar_row, column=col,
                    value=f"={mcols[m - 1]}{ar_row}+{mcols[m]}{bill_row}-{mcols[m]}{coll_row}"
                    ).number_format = "#,##0"
    _note(ws.cell(row=ar_row, column=2),
          "Receivables = what has been billed but not yet collected. Peak AR (below) is the "
          "largest this gets — the most cash tied up waiting to be paid.")
    r = ar_row + 1
    ws.cell(row=r, column=1, value="Peak AR balance").font = _FONT_BOLD
    peak = ws.cell(row=r, column=2,
                   value=f"=MAX({mcols[0]}{ar_row}:{mcols[-1]}{ar_row})")
    peak.number_format = "#,##0"
    peak.font = _FONT_BOLD
    _note(peak, "MAX of the monthly AR row — the peak receivables exposure. Live: scales with "
                "units, ASP and Demand%.")



# ═══════════════════════════════════════════════════════════════════════════════
# Variance, Assumption Register, CRB Summary, Changelog
# ═══════════════════════════════════════════════════════════════════════════════

def _write_variance(ws, versions, pkg: DealPackage, as_of: str) -> None:
    ws.title = "Variance"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80
    _title_block(ws, 2, pkg.deal_name, "Variance bridge", as_of)
    if versions and len(versions) >= 2:
        ws.cell(row=2, column=1, value="VARIANCE BRIDGE").font = _FONT_BOLD
        for col, h in enumerate(["Driver", "Δ on gross margin (USD)"], 1):
            ws.cell(row=3, column=col, value=h).font = _FONT_BOLD
        # (Populated from engine variance output when ≥2 versions exist.)
    else:
        _wrapped(ws, 3, 1,
                 "No variance to show yet. The variance bridge walks driver-by-driver between two "
                 "saved deal versions (e.g. 'Counterparty initial' → 'Our counter v1'), attributing "
                 "each term/assumption change to its dollar impact on net revenue, gross margin and "
                 "NPV, summing exactly to the total delta. It populates automatically once a second "
                 "version is recorded — there is only one version on this deal today.", height=90)
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)


def _write_assumption_reg(ws, register, pkg: DealPackage, as_of: str) -> None:
    ws.title = "Assumption_Reg"
    for col, w in {"A": 34, "B": 26, "C": 18, "D": 20, "E": 22, "F": 28, "G": 46}.items():
        ws.column_dimensions[col].width = w
    _title_block(ws, 7, pkg.deal_name, "Assumption register — what each input is & who owns it", as_of)
    for col, h in enumerate(["Field path", "Label", "Value", "Type", "Provenance", "Owner", "Note"], 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = _FONT_BOLD
        c.fill = _F_SECTION_HDR
    ws.freeze_panes = "A3"
    r = 3
    for e in register:
        ws.cell(row=r, column=1, value=e.field_path).alignment = _WRAP
        ws.cell(row=r, column=2, value=e.label).alignment = _WRAP
        cval = ws.cell(row=r, column=3, value=str(e.value) if e.value is not None else "")
        fill = _ATYPE_FILLS.get(e.assumption_type)
        if fill:
            cval.fill = fill
        ws.cell(row=r, column=4, value=e.assumption_type.value if e.assumption_type else "")
        ws.cell(row=r, column=5, value=e.basis.value if e.basis else "")
        ws.cell(row=r, column=6, value=e.owner).alignment = _WRAP
        _wrapped(ws, r, 7, e.note or "")
        r += 1


def _write_crb_summary(ws, econ, warrant, memo, pkg: DealPackage, as_of: str, cell_map: dict) -> None:
    ws.title = "CRB_Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 95
    _title_block(ws, 2, pkg.deal_name, "Contract Review Board — one-page summary", as_of)

    r = 2
    for label, val in (("Counterparty", pkg.counterparty or ""),
                       ("Status", pkg.status.value if pkg.status else ""),
                       ("Archetype", pkg.archetype or "")):
        ws.cell(row=r, column=1, value=label).font = _FONT_BOLD
        ws.cell(row=r, column=2, value=val); r += 1
    r += 1

    _section_hdr(ws, r, "ECONOMICS (live — references the Model tab)", 2); r += 1
    econ_rows = [
        ("Cash net revenue", "=ModelNetCashTotal", "#,##0"),
        ("Cash gross margin", "=ModelGmCashTotal", "#,##0"),
        ("Cash gross margin %", "=ModelGmPctCashTotal", "0.0%"),
        ("GAAP net revenue", "=ModelNetGaapTotal", "#,##0"),
        ("GAAP gross margin", "=ModelGmGaapTotal", "#,##0"),
        ("NPV (pre-tax, live)", "=ModelNPV", "#,##0"),
        ("All-in net ASP / unit", "=ASP-ModelActiveRebateTotal/TotalUnits-WarrantTotalEFV/TotalUnits", "#,##0.00"),
    ]
    for label, f, fmt in econ_rows:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=f).number_format = fmt
        r += 1
    r += 1

    _section_hdr(ws, r, "GAAP ↔ CASH BRIDGE", 2); r += 1
    _wrapped(ws, r, 2,
             "Cash margin ~37.6% is healthy; GAAP margin is negative because the $3.38B warrant "
             "(free stock to the customer) exceeds the $1.36B product margin and is subtracted as "
             "contra-revenue. The deal is cash-positive but GAAP-dilutive — AMD pays in equity, "
             "not cash.", height=70); r += 2

    if warrant:
        _section_hdr(ws, r, "WARRANT", 2); r += 1
        ws.cell(row=r, column=1, value="Total expected fair value")
        ws.cell(row=r, column=2, value="=WarrantTotalEFV").number_format = "#,##0"; r += 1
        _wrapped(ws, r, 2,
                 "Correlation caveat (§4): the spot-price + independent-vest-probability valuation "
                 "likely understates upside-scenario warrant cost, because deployment milestones "
                 "and stock-price hurdles are positively correlated.", height=55); r += 2

    if memo:
        _section_hdr(ws, r, "POLICY VERDICT", 2); r += 1
        ws.cell(row=r, column=1, value="Verdict")
        _static(ws.cell(row=r, column=2), str(memo.policy_verdict.overall_outcome.value)
                if memo.policy_verdict else "(see policy engine)"); r += 1
        ws.cell(row=r, column=1, value="Required approvers")
        approvers = []
        if memo.policy_verdict and memo.policy_verdict.all_required_approvers:
            approvers = [str(a) for a in memo.policy_verdict.all_required_approvers]
        _static(ws.cell(row=r, column=2), ", ".join(approvers)); r += 2

        _section_hdr(ws, r, "TOP RISKS", 2); r += 1
        for risk in (memo.top_risks or [])[:5]:
            desc = getattr(risk, "description", None) or getattr(risk, "title", str(risk))
            _wrapped(ws, r, 2, f"• {desc}"); r += 1
        r += 1

        if memo.benchmark_sentences:
            _section_hdr(ws, r, "BENCHMARKS", 2); r += 1
            for s in memo.benchmark_sentences:
                _wrapped(ws, r, 2, f"• {s}"); r += 1
            r += 1

        if memo.gap_report_lines:
            _section_hdr(ws, r, "TOP ASSUMPTION GAPS", 2); r += 1
            for g in memo.gap_report_lines[:5]:
                q = getattr(g, "question", str(g))
                _wrapped(ws, r, 2, f"• {q}"); r += 1
            r += 1

        if memo.approval_conditions:
            _section_hdr(ws, r, "APPROVAL CONDITIONS", 2); r += 1
            for c in memo.approval_conditions:
                _wrapped(ws, r, 2, f"• {c}"); r += 1
            r += 1

        _section_hdr(ws, r, "RECOMMENDATION", 2); r += 1
        _wrapped(ws, r, 2, memo.recommendation or "", height=70)


def _write_changelog(ws, pkg: DealPackage, as_of: str) -> None:
    ws.title = "Changelog"
    for col, w in {"A": 22, "B": 30, "C": 20, "D": 20, "E": 36, "F": 18}.items():
        ws.column_dimensions[col].width = w
    _title_block(ws, 6, pkg.deal_name, "Change journal (audit trail)", as_of)
    journal = pkg.change_journal or []
    if not journal:
        _wrapped(ws, 2, 1,
                 "No edits recorded yet. Every assumption or term change on this deal will be "
                 "logged here (timestamp, field, old value, new value, note, who) and feeds the "
                 "variance-bridge narrative. This deal is at its initial build, so the journal is "
                 "empty by design.", height=70)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        return
    for col, h in enumerate(["Timestamp", "Field", "Old", "New", "Note", "Actor"], 1):
        ws.cell(row=2, column=col, value=h).font = _FONT_BOLD
    r = 3
    for e in journal:
        ws.cell(row=r, column=1, value=str(e.timestamp))
        ws.cell(row=r, column=2, value=e.field_path)
        ws.cell(row=r, column=3, value=str(e.old_value) if e.old_value is not None else "")
        ws.cell(row=r, column=4, value=str(e.new_value) if e.new_value is not None else "")
        _wrapped(ws, r, 5, e.note or "")
        ws.cell(row=r, column=6, value=e.actor)
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# Analysis tabs (narrative, live numbers via TEXT())
# ═══════════════════════════════════════════════════════════════════════════════

def _t_usd_b(named: str) -> str:
    """Formula fragment: a named USD total formatted as $X.XXB."""
    return f'"$"&TEXT({named}/1000000000,"0.00")&"B"'


def _t_pct(named: str) -> str:
    return f'TEXT({named},"0.0%")'


def _write_analysis_fm(ws, pkg: DealPackage, as_of: str) -> None:
    ws.title = "Analysis — Finance Manager"
    ws.column_dimensions["A"].width = 115
    _title_block(ws, 1, pkg.deal_name, "Analysis — Finance Manager's read", as_of)
    rows = [
        ("HEADLINE", True),
        ('="Committed product value is "&' + _t_usd_b("ModelGrossRevTotal") +
         '&" of gross revenue across 150,000 chips. But the economics are dominated by the '
         'warrant: AMD hands Meta free stock worth "&' + _t_usd_b("WarrantTotalEFV") +
         '&", larger than the product gross margin."', False),
        ("", False),
        ("THE REAL ECONOMICS", True),
        ('="Cash / commercial gross margin is "&' + _t_usd_b("ModelGmCashTotal") + '&" ("&' +
         _t_pct("ModelGmPctCashTotal") + '&"), a healthy product margin. This is the '
         '\'non-GAAP / cash-economic\' view — it excludes the non-cash warrant, analogous to how '
         'AMD\'s own non-GAAP results exclude stock-based compensation."', False),
        ('="On a GAAP basis the warrant is contra-revenue, so GAAP net revenue collapses to "&' +
         _t_usd_b("ModelNetGaapTotal") + '&" and GAAP gross margin is "&' + _t_usd_b("ModelGmGaapTotal") +
         '&" — negative. The deal is cash-positive but GAAP-dilutive: AMD pays for it in equity, not cash."', False),
        ('="Effective all-in net ASP is $"&TEXT(ASP-ModelActiveRebateTotal/TotalUnits-WarrantTotalEFV/TotalUnits,"#,##0")'
         '&" per unit versus a $"&TEXT(ASP,"#,##0")&" sticker — the warrant and rebates erode roughly 94% of headline price."', False),
        ("", False),
        ("WHAT I'D FLAG", True),
        ('="1) Rebate ambiguity (§5): the prospective vs retroactive reading is a ~$41M swing. '
         'Default is the retroactive reading (B); resolve with Legal + Revenue Accounting before signature."', False),
        ('="2) Working capital: net-90 terms tie up real cash; peak receivables run to the hundreds of millions. '
         'The $500M prepayment cushions early quarters."', False),
        ('="3) Warrant correlation: our valuation uses independent vest odds and one spot price, so it likely '
         'UNDERSTATES warrant cost in the upside — the expensive hurdles clear exactly when the deal succeeds."', False),
        ('="Note: a real AMD hyperscaler deal would spread equity over far more volume (6+ GW), so it would not '
         'necessarily go GAAP-negative. This synthetic deal\'s warrant is simply large relative to its volume."', False),
        ("", False),
        ("RECOMMENDATION", True),
        ('="Approve subject to (a) Legal/Revenue Accounting resolving the §5 rebate reading, (b) Treasury '
         'confirming WACC and the working-capital draw, and (c) the deal team signing off on the warrant vest '
         'probabilities. The cash economics are sound; the GAAP optics must be communicated to the board up front."', False),
    ]
    r = 2
    for content, is_hdr in rows:
        c = ws.cell(row=r, column=1, value=content if content else None)
        if is_hdr:
            c.font = _FONT_BOLD
            c.fill = _F_SECTION_HDR
        else:
            c.alignment = _WRAP
            if content:
                ws.row_dimensions[r].height = 46
        r += 1


def _write_analysis_plain(ws, pkg: DealPackage, as_of: str) -> None:
    ws.title = "Analysis — Plain English"
    ws.column_dimensions["A"].width = 115
    _title_block(ws, 1, pkg.deal_name, "Analysis — Plain English (start here if you're new)", as_of)
    rows = [
        ("WHAT'S THE DEAL?", True),
        ('="AMD sells Meta 150,000 AI computer chips at $25,000 each — about "&' + _t_usd_b("ModelGrossRevTotal") +
         '&" of chips over three years."', False),
        ("", False),
        ("WHERE'S THE CATCH?", True),
        ('="AMD ALSO gives Meta free stock (a \'warrant\') worth about "&' + _t_usd_b("WarrantTotalEFV") +
         '&". That free stock is worth MORE than the profit AMD makes on the chips themselves."', False),
        ('="So two scoreboards disagree. Counting only cash, AMD earns a solid profit: about "&' +
         _t_usd_b("ModelGmCashTotal") + '&" of margin ("&' + _t_pct("ModelGmPctCashTotal") +
         '&"). But the official accounting (GAAP) subtracts the free stock, and on that scoreboard the '
         'deal LOSES money: "&' + _t_usd_b("ModelGmGaapTotal") + '&"."', False),
        ('="Think of it like AMD paying a giant bonus in company shares instead of cash. The cash register '
         'looks great; the accountant\'s books look bad. (Tech companies often show a \'non-GAAP\' number that '
         'leaves out share-based pay for exactly this reason.)"', False),
        ("", False),
        ("WHAT ARE THE RISKS?", True),
        ('="1) A fuzzy sentence about discounts could cost ~$41M depending on how lawyers read it."', False),
        ('="2) Meta gets 90 days to pay, so AMD waits on a lot of cash (a $500M upfront payment helps)."', False),
        ('="3) The free stock gets MORE expensive if AMD\'s share price climbs — and it climbs when the deal '
         'goes well. Success has a cost here."', False),
        ('="(In a bigger real-world deal AMD would spread that free stock over far more chips, so it wouldn\'t '
         'swamp the profit like it does in this small synthetic example.)"', False),
        ("", False),
        ("IS IT A GOOD DEAL?", True),
        ('="In cash terms, yes — it makes money and locks in a huge customer. The warning label: on paper '
         '(GAAP) it looks like a loss because of the free stock, so leadership needs to explain that up front. '
         'Settle the discount wording with Legal before signing."', False),
    ]
    r = 2
    for content, is_hdr in rows:
        c = ws.cell(row=r, column=1, value=content if content else None)
        if is_hdr:
            c.font = _FONT_BOLD
            c.fill = _F_SECTION_HDR
        else:
            c.alignment = _WRAP
            if content:
                ws.row_dimensions[r].height = 46
        r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# Named ranges + public API
# ═══════════════════════════════════════════════════════════════════════════════

def _add_named_ranges(wb: Workbook, cell_map: dict) -> None:
    for name, addr in cell_map.items():
        if not isinstance(addr, str) or "!" not in addr:
            continue
        try:
            wb.defined_names[name] = DefinedName(name, attr_text=addr)
        except Exception:
            pass


def build_workbook(
    pkg: DealPackage,
    assumptions: DealAssumptions,
    econ: DealEconomics,
    warrant: WarrantEconomics | None,
    register: list[RegisterEntry],
    memo: CRBMemo | None = None,
    versions: list[DealVersion] | None = None,
    as_of: str | None = None,
) -> Workbook:
    """Build the 14-tab, fully-live, self-documenting deal-model workbook.

    Every model input is an editable named cell; every downstream number is a live
    formula. The only static numeric cells are the labeled input pins and the
    Scenarios tab's DOWNSIDE/UPSIDE/ET engine snapshots (loudly banner-labeled).
    Returns the Workbook; the caller saves it.
    """
    from deal_copilot.economics_engine import extract_inputs

    inp = extract_inputs(pkg, assumptions=assumptions)
    n_q = len(inp.committed_quarterly)
    as_of = as_of or datetime.now().strftime("%Y-%m-%d")
    # Model row layout is generated from the actual tier count, and shared with
    # every tab that references a Model row.
    rows = _ModelRows(len(inp.rebate_tiers))

    wb = Workbook()
    wb.remove(wb.active)
    cell_map: dict[str, Any] = {}

    # Pre-commit the Warrant contra-per-quarter addresses so the Model can read
    # them before the Warrant tab is physically written.
    for q in range(n_q):
        cell_map[f"ContraSchedQ{q + 1}"] = f"Warrant!${_cl(_Q1_COL + q)}${_W_CONTRA_ROW}"

    _write_cover(wb.create_sheet("Cover"), pkg, as_of)
    _write_assumptions(wb.create_sheet("Assumptions"), assumptions, inp, pkg, as_of, cell_map)
    _write_warrant_assump(wb.create_sheet("Warrant_Assump"), assumptions, warrant, pkg, as_of, cell_map)
    ws_drv = wb.create_sheet("Drivers")
    _write_drivers(ws_drv, econ, pkg, inp, as_of, cell_map)
    _write_tier_table(ws_drv, inp, cell_map)
    _write_model(wb.create_sheet("Model"), inp, pkg, as_of, cell_map, rows)
    _write_warrant(wb.create_sheet("Warrant"), warrant, inp, pkg, as_of, cell_map, rows)
    _write_scenarios(wb.create_sheet("Scenarios"), econ, pkg, as_of, cell_map)
    _write_acct_sched(wb.create_sheet("Acct_Sched"), inp, pkg, as_of, cell_map, rows)
    _write_variance(wb.create_sheet("Variance"), versions, pkg, as_of)
    _write_assumption_reg(wb.create_sheet("Assumption_Reg"), register, pkg, as_of)
    _write_crb_summary(wb.create_sheet("CRB_Summary"), econ, warrant, memo, pkg, as_of, cell_map)
    _write_changelog(wb.create_sheet("Changelog"), pkg, as_of)
    _write_analysis_fm(wb.create_sheet("Analysis — Finance Manager"), pkg, as_of)
    _write_analysis_plain(wb.create_sheet("Analysis — Plain English"), pkg, as_of)

    _add_named_ranges(wb, cell_map)
    return wb


__all__ = ["build_workbook"]
